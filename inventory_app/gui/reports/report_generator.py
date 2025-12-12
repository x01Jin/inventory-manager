"""
Report generator for the inventory application.
Handles Excel report generation for dynamic reports based on date range.
Refactored to use QueryBuilder for SQL operations and eliminate redundancy.
"""

from typing import List, Dict, Optional
from datetime import date, datetime, timedelta
from pathlib import Path

from inventory_app.database.connection import db
from inventory_app.utils.logger import logger
from inventory_app.gui.reports.report_utils import date_formatter
from inventory_app.gui.reports.query_builder import (
    ReportQueryBuilder,
    ReportStatisticsBuilder,
)
from inventory_app.services.movement_types import MovementType

# Import openpyxl directly since it's in requirements.txt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """
    Handles report generation and Excel export.
    Uses composition with DatabaseConnection and openpyxl.
    """

    def __init__(self):
        """Initialize report generator."""

    def get_granularity(self, start_date: date, end_date: date) -> str:
        """
        Determine the appropriate reporting granularity based on date range.

        Args:
            start_date (datetime.date): Start date of the reporting period
            end_date (datetime.date): End date of the reporting period

        Returns:
            str: Granularity level ('daily', 'weekly', 'monthly', 'quarterly', 'yearly', 'multi_year')
        """
        return date_formatter.get_smart_granularity(start_date, end_date)

    def generate_report(
        self,
        start_date: date,
        end_date: date,
        output_path: Optional[str] = None,
        grade_filter: str = "",
        section_filter: str = "",
        include_consumables: bool = True,
    ) -> str:
        """
        Generate dynamic usage report in Excel format based on date range.

        Args:
            start_date: Start date for the report period
            end_date: End date for the report period
            output_path: Optional output file path
            grade_filter: Filter by grade level
            section_filter: Filter by section
            include_consumables: Whether to include consumable items

        Returns:
            Path to generated Excel file
        """
        # Determine granularity and report title
        granularity = self.get_granularity(start_date, end_date)

        try:
            logger.info(
                f"Generating {granularity} report from {start_date} to {end_date}"
            )

            # Get report data with dynamic granularity
            report_data = self._get_dynamic_report_data(
                start_date,
                end_date,
                granularity,
                grade_filter,
                section_filter,
                include_consumables,
            )

            if not report_data:
                error_msg = "Failed to generate report\nReason: No data found for the specified period"
                logger.warning("No data found for the specified period")
                return error_msg

            # Create Excel file
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"{granularity}_report_{timestamp}.xlsx"

            output_path_obj = Path(output_path)
            title = f"{granularity.title()} Usage Report"
            self._create_excel_report(
                report_data, output_path_obj, title, start_date, end_date
            )

            logger.info(f"{granularity.title()} report generated: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Failed to generate {granularity} report: {e}")
            return ""

    def _get_dynamic_report_data(
        self,
        start_date: date,
        end_date: date,
        granularity: str,
        grade_filter: str = "",
        section_filter: str = "",
        include_consumables: bool = True,
    ) -> List[Dict]:
        """
        Get dynamic report data from database using QueryBuilder.

        Args:
            start_date: Start date for the report
            end_date: End date for the report
            granularity: 'daily', 'weekly', 'monthly', or 'quarterly'
            grade_filter: Filter by grade level
            section_filter: Filter by section
            include_consumables: Whether to include consumable items

        Returns:
            List of report rows with dynamic period columns
        """
        try:
            # Use QueryBuilder for SQL generation and execution
            query_builder = ReportQueryBuilder()

            # Build the dynamic query (optimized approach with built-in period columns)
            query, params = query_builder.build_dynamic_report_query(
                start_date,
                end_date,
                granularity,
                grade_filter,
                section_filter,
                include_consumables,
            )

            # Execute query and process results; pivot if normalized fallback was used
            rows = query_builder.execute_report_query(query, params)
            if not rows:
                return []

            # If query builder used normalized fallback, rows will include
            # PERIOD and PERIOD_QUANTITY columns that we need to pivot
            if query_builder.normalized_fallback or (
                rows and "PERIOD" in rows[0] and "PERIOD_QUANTITY" in rows[0]
            ):
                # Build pivoted structure keyed by item identity
                from collections import OrderedDict

                # Fetch period keys in the same order as the formatter
                period_keys = date_formatter.get_period_keys(
                    start_date, end_date, granularity
                )
                pivoted = OrderedDict()
                for row in rows:
                    # Build a stable item key from the fixed columns
                    item_key = (
                        row.get("ITEMS"),
                        row.get("CATEGORIES"),
                        row.get("ACTUAL_INVENTORY"),
                        row.get("SIZE"),
                        row.get("BRAND"),
                        row.get("OTHER SPECIFICATIONS"),
                    )
                    if item_key not in pivoted:
                        # Initialize base structure with zeros for all periods
                        base = {
                            "ITEMS": item_key[0],
                            "CATEGORIES": item_key[1],
                            "ACTUAL_INVENTORY": item_key[2],
                            "SIZE": item_key[3],
                            "BRAND": item_key[4],
                            "OTHER SPECIFICATIONS": item_key[5],
                        }
                        for k in period_keys:
                            base[k] = 0
                        pivoted[item_key] = base

                    # Map period_quantity into the pivot map
                    pkey = row.get("PERIOD")
                    qty = row.get("PERIOD_QUANTITY") or 0
                    if pkey in pivoted[item_key]:
                        pivoted[item_key][pkey] += qty
                    else:
                        pivoted[item_key][pkey] = qty

                # Convert pivoted OrderedDict values into a list and compute total quantity
                result_rows = []
                for base in pivoted.values():
                    total = 0
                    for k in period_keys:
                        total += base.get(k, 0) or 0
                    base["TOTAL QUANTITY"] = total
                    result_rows.append(base)
                return result_rows

            return rows

        except Exception as e:
            logger.error(f"Failed to get dynamic report data: {e}")
            return []

    def _format_excel_headers(
        self, headers: List[str], start_date: date, end_date: date
    ) -> List[str]:
        """
        Format Excel headers by converting period keys to user-friendly format.

        Args:
            headers: Raw headers from database
            start_date: Report start date
            end_date: Report end date

        Returns:
            List of formatted headers
        """
        try:
            # Determine granularity for the report
            granularity = self.get_granularity(start_date, end_date)
            formatted_headers = []

            # Fixed headers that don't need formatting
            fixed_headers = {
                "ITEMS",
                "CATEGORIES",
                "ACTUAL_INVENTORY",
                "SIZE",
                "BRAND",
                "OTHER SPECIFICATIONS",
                "TOTAL QUANTITY",
            }

            for header in headers:
                if header in fixed_headers:
                    # Keep fixed headers as-is
                    formatted_headers.append(header)
                else:
                    # Try to parse as period key and format it
                    try:
                        formatted_header = self._parse_and_format_period_key(
                            header, granularity
                        )
                        formatted_headers.append(formatted_header)
                    except Exception:
                        # If parsing fails, keep original header
                        logger.warning(f"Could not parse period key: {header}")
                        formatted_headers.append(header)

            return formatted_headers

        except Exception as e:
            logger.error(f"Failed to format Excel headers: {e}")
            return headers  # Return original headers if formatting fails

    def _parse_and_format_period_key(self, period_key: str, granularity: str) -> str:
        """
        Parse a period key and format it according to granularity, including excess periods.

        Args:
            period_key: Raw period key from database (e.g., '2023-01-01', '2023-12-01to02', '2023-Q1')
            granularity: Report granularity

        Returns:
            Formatted period header string
        """
        try:
            # Handle excess period keys first (format: YYYY-MM-DDtoDD)
            if "to" in period_key:
                start_date_str, end_date_str = period_key.split("to")
                start_date = date.fromisoformat(start_date_str)
                end_date = date.fromisoformat(end_date_str)

                # Format as (Month/DayRange/Year)
                month_names = [
                    "Jan",
                    "Feb",
                    "Mar",
                    "Apr",
                    "May",
                    "Jun",
                    "Jul",
                    "Aug",
                    "Sep",
                    "Oct",
                    "Nov",
                    "Dec",
                ]

                start_month = month_names[start_date.month - 1]
                end_month = month_names[end_date.month - 1]

                if start_date.month == end_date.month:
                    # Same month: (Dec/01-02/2024)
                    return f"({start_month}/{start_date.day:02d}-{end_date.day:02d}/{start_date.year})"
                else:
                    # Different months: (Nov/29-Dec/02/2024)
                    return f"({start_month}/{start_date.day:02d}-{end_month}/{end_date.day:02d}/{start_date.year})"

            # Parse different period key formats
            if granularity == "daily":
                # Format: '2023-01-01'
                parsed_date = date.fromisoformat(period_key)
                return date_formatter.format_period_header(parsed_date, "daily")

            elif granularity == "weekly":
                # Format: '2023-01-W1' (year-month-week)
                try:
                    year_str, month_str, week_part = period_key.split("-")
                    year, month = int(year_str), int(month_str)
                    week_str = week_part.replace("W", "")
                    week_num = int(week_str)

                    # Calculate a representative date for this month-week
                    # Use the first day of the week within the month
                    first_day_of_month = date(year, month, 1)
                    week_date = first_day_of_month + timedelta(days=(week_num - 1) * 7)

                    return date_formatter.format_period_header(week_date, "weekly")
                except (ValueError, IndexError):
                    # Fallback if parsing fails
                    return period_key

            elif granularity == "monthly":
                # Format: '2023-01'
                year_str, month_str = period_key.split("-")
                year, month = int(year_str), int(month_str)
                month_date = date(year, month, 1)

                return date_formatter.format_period_header(month_date, "monthly")

            elif granularity == "quarterly":
                # Format: '2023-Q1'
                year_str, quarter_str = period_key.split("-Q")
                year, quarter = int(year_str), int(quarter_str)

                # Calculate first month of quarter
                quarter_month = (quarter - 1) * 3 + 1
                quarter_date = date(year, quarter_month, 1)

                return date_formatter.format_period_header(quarter_date, "quarterly")

            elif granularity in ["yearly", "multi_year"]:
                # Format: '2023'
                year = int(period_key)
                year_date = date(year, 1, 1)

                return date_formatter.format_period_header(year_date, "yearly")

            else:
                # Unknown granularity, return as-is
                return period_key

        except Exception as e:
            logger.error(f"Failed to parse period key '{period_key}': {e}")
            return period_key

    def _create_excel_report(
        self,
        data: List[Dict],
        output_path: Path,
        title: str,
        start_date: date,
        end_date: date,
    ) -> None:
        """
        Create Excel file with report data.

        Args:
            data: Report data rows
            output_path: Output file path
            title: Report title
            start_date: Report start date
            end_date: Report end date
        """
        try:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("Could not create worksheet")

            ws.title = "Report"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            center_align = Alignment(horizontal="center")

            # Add title
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, size=16)

            # Use enhanced date formatting for period display
            period_description = date_formatter.get_date_range_description(
                start_date, end_date
            )
            ws["A2"] = f"Period: {period_description}"
            ws["A2"].font = Font(italic=True)

            # Add headers
            if data:
                headers = list(data[0].keys())
                formatted_headers = self._format_excel_headers(
                    headers, start_date, end_date
                )

                for col_num, header_text in enumerate(formatted_headers, 1):
                    cell = ws.cell(row=4, column=col_num)
                    # Check if this is a merged cell to avoid type errors
                    from openpyxl.cell import MergedCell

                    if not isinstance(cell, MergedCell):
                        cell.value = header_text
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = border

                    # Auto-adjust column width (always do this)
                    column_letter = get_column_letter(col_num)
                    ws.column_dimensions[column_letter].width = max(
                        len(str(header_text)) + 2, 12
                    )

            # Add data rows
            for row_num, row_data in enumerate(data, 5):
                for col_num, value in enumerate(row_data.values(), 1):
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        # Check if this is a merged cell to avoid type errors
                        from openpyxl.cell import MergedCell

                        if not isinstance(cell, MergedCell):
                            cell.value = value
                            cell.border = border

                        # Auto-adjust column width based on content (always do this)
                        column_letter = get_column_letter(col_num)
                        content_length = len(str(value)) + 2
                        current_width = ws.column_dimensions[column_letter].width or 12
                        ws.column_dimensions[column_letter].width = max(
                            current_width, content_length
                        )
                    except Exception as e:
                        # Skip cells that can't be written to (e.g., merged cells)
                        logger.debug(
                            f"Skipping cell at row {row_num}, column {col_num}: {e}"
                        )

            # Save the workbook
            wb.save(output_path)
            logger.debug(f"Excel file saved to {output_path}")

        except Exception as e:
            logger.error(f"Failed to create Excel report: {e}")
            raise

    def get_usage_statistics(self, start_date: date, end_date: date) -> Dict:
        """
        Get usage statistics for a date range using ReportStatisticsBuilder.

        Args:
            start_date: Start date
            end_date: End date

        Returns:
            Dictionary with usage statistics
        """
        try:
            stats_builder = ReportStatisticsBuilder()

            # Get total items used
            total_query, total_params = stats_builder.build_usage_statistics_query(
                start_date, end_date
            )
            total_rows = db.execute_query(total_query, total_params)
            total_used = (
                total_rows[0]["total_used"]
                if total_rows and total_rows[0]["total_used"]
                else 0
            )

            # Get items by category
            category_query, category_params = (
                stats_builder.build_category_statistics_query(start_date, end_date)
            )
            category_rows = db.execute_query(category_query, category_params) or []

            # Get top used items
            top_items_query, top_items_params = stats_builder.build_top_items_query(
                start_date, end_date
            )
            top_items_rows = db.execute_query(top_items_query, top_items_params) or []

            return {
                "total_items_used": total_used,
                "categories": category_rows,
                "top_items": top_items_rows,
                "date_range": f"{start_date} to {end_date}",
            }

        except Exception as e:
            logger.error(f"Failed to get usage statistics: {e}")
            return {
                "total_items_used": 0,
                "categories": [],
                "top_items": [],
                "date_range": f"{start_date} to {end_date}",
            }

    def generate_inventory_report(
        self,
        report_type: str,
        start_date: date,
        end_date: date,
        category_filter: str = "",
        output_path: Optional[str] = None,
    ) -> str:
        """
        Generate inventory report based on type.

        Args:
            report_type: Type of inventory report ('stock_levels', 'expiration', 'low_stock', 'acquisition_history', 'calibration_due')
            start_date: Start date for the report period
            end_date: End date for the report period
            category_filter: Category filter
            output_path: Optional output file path

        Returns:
            Path to generated Excel file
        """
        try:
            logger.info(
                f"Generating {report_type} report from {start_date} to {end_date}"
            )

            # Get report data based on type
            if report_type == "Stock Levels Report":
                report_data = self._get_stock_levels_data(category_filter)
                title = "Stock Levels Report"
            elif report_type == "Expiration Report":
                report_data = self._get_expiration_data(
                    start_date, end_date, category_filter
                )
                title = "Expiration Report"
            elif report_type == "Low Stock Alert":
                report_data = self._get_low_stock_data(category_filter)
                title = "Low Stock Alert Report"
            elif report_type == "Acquisition History":
                report_data = self._get_acquisition_history_data(
                    start_date, end_date, category_filter
                )
                title = "Acquisition History Report"
            elif report_type == "Calibration Due Report":
                report_data = self._get_calibration_due_data(
                    start_date, end_date, category_filter
                )
                title = "Calibration Due Report"
            else:
                return f"Unknown inventory report type: {report_type}"

            if not report_data:
                error_msg = f"Failed to generate {report_type}\nReason: No data found for the specified criteria"
                logger.warning(f"No data found for {report_type}")
                return error_msg

            # Create Excel file
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"inventory_{report_type.lower().replace(' ', '_')}_{timestamp}.xlsx"

            output_path_obj = Path(output_path)
            self._create_excel_report(
                report_data, output_path_obj, title, start_date, end_date
            )

            logger.info(f"{report_type} report generated: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Failed to generate {report_type} report: {e}")
            return ""

    def _get_stock_levels_data(self, category_filter: str = "") -> List[Dict]:
        """Get current stock levels data.

        Parameters:
            category_filter: Optional category filter
            min_stock: If provided, return only items with current stock < min_stock
        """
        try:
            query = """
            SELECT
                i.name AS "Item Name",
                c.name AS "Category",
                i.size AS "Size",
                i.brand AS "Brand",
                COALESCE(SUM(ib.quantity_received), 0) - COALESCE(SUM(sm.quantity), 0) AS "Current Stock",
                i.other_specifications AS "Specifications"
            FROM Items i
            JOIN Categories c ON c.id = i.category_id
            LEFT JOIN Item_Batches ib ON ib.item_id = i.id AND ib.disposal_date IS NULL
            LEFT JOIN Stock_Movements sm ON sm.item_id = i.id AND sm.movement_type IN ('%s', '%s')
            """

            params = []
            if category_filter:
                query += " WHERE c.name = ?"
                params.append(category_filter)

            query += " GROUP BY i.id, i.name, c.name, i.size, i.brand, i.other_specifications ORDER BY c.name, i.name"
            query = query % (
                MovementType.CONSUMPTION.value,
                MovementType.DISPOSAL.value,
            )

            return db.execute_query(query, tuple(params)) or []

        except Exception as e:
            logger.error(f"Failed to get stock levels data: {e}")
            return []

    def _get_expiration_data(
        self, start_date: date, end_date: date, category_filter: str = ""
    ) -> List[Dict]:
        """Get items expiring within date range."""
        try:
            query = """
            SELECT
                i.name AS "Item Name",
                c.name AS "Category",
                i.size AS "Size",
                i.brand AS "Brand",
                i.expiration_date AS "Expiration Date",
                COALESCE(SUM(ib.quantity_received), 0) AS "Stock Quantity",
                i.other_specifications AS "Specifications"
            FROM Items i
            JOIN Categories c ON c.id = i.category_id
            LEFT JOIN Item_Batches ib ON ib.item_id = i.id AND ib.disposal_date IS NULL
            WHERE i.expiration_date BETWEEN ? AND ?
            """

            params = [start_date.isoformat(), end_date.isoformat()]

            if category_filter:
                query += " AND c.name = ?"
                params.append(category_filter)

            query += " GROUP BY i.id, i.name, c.name, i.size, i.brand, i.expiration_date, i.other_specifications ORDER BY i.expiration_date"

            return db.execute_query(query, tuple(params)) or []

        except Exception as e:
            logger.error(f"Failed to get expiration data: {e}")
            return []

    def _get_low_stock_data(
        self, category_filter: str = "", threshold: int = 10
    ) -> List[Dict]:
        """Get items with low stock (< threshold). Re-uses stock levels query to avoid duplication."""
        try:
            rows = self._get_stock_levels_data(category_filter=category_filter)
            if not rows:
                return []

            filtered = [r for r in rows if (r.get("Current Stock") or 0) < threshold]
            filtered = sorted(filtered, key=lambda r: (r.get("Current Stock") or 0))
            return filtered

        except Exception as e:
            logger.error(f"Failed to get low stock data: {e}")
            return []

    def _get_acquisition_history_data(
        self, start_date: date, end_date: date, category_filter: str = ""
    ) -> List[Dict]:
        """Get acquisition history within date range."""
        try:
            query = """
            SELECT
                i.name AS "Item Name",
                c.name AS "Category",
                ib.date_received AS "Acquisition Date",
                ib.quantity_received AS "Quantity Received",
                ib.batch_number AS "Batch Number",
                s.name AS "Supplier",
                i.other_specifications AS "Specifications"
            FROM Item_Batches ib
            JOIN Items i ON i.id = ib.item_id
            JOIN Categories c ON c.id = i.category_id
            LEFT JOIN Suppliers s ON s.id = i.supplier_id
            WHERE ib.date_received BETWEEN ? AND ?
            """

            params = [start_date.isoformat(), end_date.isoformat()]

            if category_filter:
                query += " AND c.name = ?"
                params.append(category_filter)

            query += " ORDER BY ib.date_received DESC, i.name"

            return db.execute_query(query, tuple(params)) or []

        except Exception as e:
            logger.error(f"Failed to get acquisition history data: {e}")
            return []

    def _get_calibration_due_data(
        self, start_date: date, end_date: date, category_filter: str = ""
    ) -> List[Dict]:
        """Get items needing calibration within date range."""
        try:
            query = """
            SELECT
                i.name AS "Item Name",
                c.name AS "Category",
                i.size AS "Size",
                i.brand AS "Brand",
                i.calibration_date AS "Calibration Date",
                i.other_specifications AS "Specifications"
            FROM Items i
            JOIN Categories c ON c.id = i.category_id
            WHERE i.calibration_date BETWEEN ? AND ?
            """

            params = [start_date.isoformat(), end_date.isoformat()]

            if category_filter:
                query += " AND c.name = ?"
                params.append(category_filter)

            query += " ORDER BY i.calibration_date"

            return db.execute_query(query, tuple(params)) or []

        except Exception as e:
            logger.error(f"Failed to get calibration due data: {e}")
            return []


# Global report generator instance
report_generator = ReportGenerator()

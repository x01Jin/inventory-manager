"""
Monthly Usage Report Generator for Laboratory Inventory.

Generates Excel reports in the exact format specified in beta test requirements:
- Row 1: Empty (spacing)
- Row 2: Header merged in columns F-G ("REPORT ON THE USAGE OF LABORATORY MATERIALS...")
- Row 3: Empty (spacing)
- Row 4-6: Column headers with month name and week date ranges
- Columns A-F: Item details (ITEMS, CATEGORIES, ACTUAL INVENTORY, SIZE, BRAND, OTHER SPECS)
- Columns G+: Week columns (PRE, WEEK 1-4, POST) with date ranges, Total

Week structure includes pre/post excess days:
- PRE: Days before first full week (if month doesn't start on Monday)
- WEEK 1-4: Full weeks within the month
- POST: Days after last full week (if month doesn't end on Sunday)

Per beta test requirements #20, #21, #17:
- Report format matches sample Excel files
- Items classified by exact category
- Weekly breakdown within months
"""

from typing import List, Dict, Optional, Tuple, NamedTuple
from datetime import date, timedelta
from calendar import monthrange
from pathlib import Path
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from inventory_app.database.connection import db
from inventory_app.utils.logger import logger


# Category display order matching sample reports
CATEGORY_ORDER = [
    "Equipment",
    "Apparatus",
    "Lab Models",
    "Chemicals-Solid",
    "Chemicals-Liquid",
    "Prepared Slides",
    "Consumables",
    "Others",
    "Uncategorized",
]

# Category display names for report headers
CATEGORY_DISPLAY_NAMES = {
    "Equipment": "EQUIPMENT",
    "Apparatus": "APPARATUSES",
    "Lab Models": "LAB MODELS",
    "Chemicals-Solid": "CHEMICALS - SOLID",
    "Chemicals-Liquid": "CHEMICALS - LIQUID",
    "Prepared Slides": "PREPARED SLIDES",
    "Consumables": "CONSUMABLES",
    "Others": "OTHERS",
    "Uncategorized": "UNCATEGORIZED",
}


class WeekPeriod(NamedTuple):
    """Represents a week period with its date range and display information."""

    start: date
    end: date
    label: str  # e.g., "WEEK 1", "PRE", "POST"
    date_range: str  # e.g., "Oct 3-7"


def _normalize_grade_level(raw_grade: str) -> str:
    """Normalize grade level text to canonical labels: Grade 7..10."""
    value = (raw_grade or "").strip()
    if not value:
        return ""

    digits = "".join(ch for ch in value if ch.isdigit())
    if digits in {"7", "8", "9", "10"}:
        return f"Grade {digits}"

    lowered = value.lower()
    if lowered in {"grade 7", "grade 8", "grade 9", "grade 10"}:
        return lowered.title()
    return ""


def get_month_weeks(year: int, month: int) -> List[WeekPeriod]:
    """
    Get weekly date ranges for a specific month with PRE/POST excess handling.

    Returns list of WeekPeriod tuples representing:
    - PRE: Days from month start to Saturday before first full week (if any)
    - WEEK 1-4: Full weeks (Monday-Sunday) within the month
    - POST: Days from Monday after last full week to month end (if any)

    Per beta test requirement: weeks displayed as PRE, WEEK 1, WEEK 2, WEEK 3, WEEK 4, POST
    """
    weeks: List[WeekPeriod] = []
    first_day = date(year, month, 1)
    _, last_day_num = monthrange(year, month)
    last_day = date(year, month, last_day_num)

    # Month name abbreviations for display
    month_names = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    month_abbrev = month_names[month - 1]

    def format_date_range(start: date, end: date) -> str:
        """Format date range as 'Mon D-D' or 'Mon D' for single day."""
        if start == end:
            return f"{month_abbrev} {start.day}"
        return f"{month_abbrev} {start.day}-{end.day}"

    # Find first Monday on or after month start
    first_monday = first_day
    while first_monday.weekday() != 0:  # 0 = Monday
        first_monday += timedelta(days=1)

    # PRE: If month doesn't start on Monday, add pre-excess period
    if first_day.weekday() != 0 and first_monday <= last_day:
        pre_end = first_monday - timedelta(days=1)
        if pre_end >= first_day:
            weeks.append(
                WeekPeriod(
                    start=first_day,
                    end=pre_end,
                    label="PRE",
                    date_range=format_date_range(first_day, pre_end),
                )
            )

    # Full weeks (WEEK 1, WEEK 2, etc.)
    current_monday = first_monday
    week_num = 1

    while current_monday <= last_day:
        week_sunday = current_monday + timedelta(days=6)

        if week_sunday <= last_day:
            # Full week within month
            weeks.append(
                WeekPeriod(
                    start=current_monday,
                    end=week_sunday,
                    label=f"WEEK {week_num}",
                    date_range=format_date_range(current_monday, week_sunday),
                )
            )
            week_num += 1
            current_monday = week_sunday + timedelta(days=1)
        else:
            # POST: Partial week at end of month
            if current_monday <= last_day:
                weeks.append(
                    WeekPeriod(
                        start=current_monday,
                        end=last_day,
                        label="POST",
                        date_range=format_date_range(current_monday, last_day),
                    )
                )
            break

    return weeks


def get_monthly_usage_data(
    year: int,
    month: int,
    category_filter: str = "",
) -> Tuple[List[Dict], List[WeekPeriod]]:
    """
    Get usage data for a specific month broken down by weeks.

    Args:
        year: Year of the report
        month: Month of the report (1-12)
        category_filter: Optional category filter

    Returns:
        Tuple of (data rows, week definitions)
    """
    try:
        weeks = get_month_weeks(year, month)
        first_day = date(year, month, 1)
        _, last_day_num = monthrange(year, month)
        last_day = date(year, month, last_day_num)

        if not weeks:
            return [], []

        # Build query to get base item data with stock.
        # ACTUAL INVENTORY follows Task 10 stock semantics.
        query = """
            SELECT
                i.id AS item_id,
                i.name AS item_name,
                c.name AS category,
                CASE
                    WHEN i.is_consumable = 1 THEN
                        COALESCE(stock.original_stock, 0) -
                        COALESCE(movements.consumed_qty, 0) -
                        COALESCE(movements.disposed_qty, 0) +
                        COALESCE(movements.returned_qty, 0)
                    ELSE
                        COALESCE(stock.original_stock, 0) -
                        COALESCE(movements.disposed_qty, 0)
                END AS actual_inventory,
                i.size,
                i.brand,
                i.other_specifications
            FROM Items i
            JOIN Categories c ON c.id = i.category_id
            LEFT JOIN (
                SELECT
                    ib.item_id,
                    COALESCE(SUM(ib.quantity_received), 0) AS original_stock
                FROM Item_Batches ib
                WHERE ib.disposal_date IS NULL
                GROUP BY ib.item_id
            ) stock ON stock.item_id = i.id
            LEFT JOIN (
                SELECT
                    sm.item_id,
                    COALESCE(SUM(CASE WHEN sm.movement_type = 'CONSUMPTION' THEN sm.quantity ELSE 0 END), 0) AS consumed_qty,
                    COALESCE(SUM(CASE WHEN sm.movement_type = 'DISPOSAL' THEN sm.quantity ELSE 0 END), 0) AS disposed_qty,
                    COALESCE(SUM(CASE WHEN sm.movement_type = 'RETURN' THEN sm.quantity ELSE 0 END), 0) AS returned_qty
                FROM Stock_Movements sm
                GROUP BY sm.item_id
            ) movements ON movements.item_id = i.id
        """

        params: List = []
        if category_filter:
            query += " WHERE c.name = ?"
            params.append(category_filter)

        query += " GROUP BY i.id, i.name, c.name, i.size, i.brand, i.other_specifications, i.is_consumable, stock.original_stock, movements.consumed_qty, movements.disposed_qty, movements.returned_qty ORDER BY c.name, i.name"

        base_items = db.execute_query(query, tuple(params)) or []

        if not base_items:
            return [], weeks

        grade_usage_rows = (
            db.execute_query(
                """
                SELECT
                    ri.item_id,
                    req.grade_level,
                    SUM(ri.quantity_requested) AS quantity
                FROM Requisition_Items ri
                JOIN Requisitions r ON r.id = ri.requisition_id
                JOIN Requesters req ON req.id = r.requester_id
                WHERE r.lab_activity_date >= ? AND r.lab_activity_date <= ?
                GROUP BY ri.item_id, req.grade_level
                """,
                (first_day.isoformat(), last_day.isoformat()),
            )
            or []
        )

        grade_usage_map: Dict[int, Dict[str, int]] = {}
        for grade_row in grade_usage_rows:
            item_id = grade_row.get("item_id")
            if item_id is None:
                continue
            normalized_grade = _normalize_grade_level(
                grade_row.get("grade_level") or ""
            )
            if not normalized_grade:
                continue
            if item_id not in grade_usage_map:
                grade_usage_map[item_id] = {
                    "Grade 7": 0,
                    "Grade 8": 0,
                    "Grade 9": 0,
                    "Grade 10": 0,
                }
            grade_usage_map[item_id][normalized_grade] += int(
                grade_row.get("quantity") or 0
            )

        # Get usage per week for each item
        # NOTE: Usage based on lab_activity_date per beta test requirements
        usage_query = """
            SELECT
                ri.item_id,
                SUM(ri.quantity_requested) AS quantity
            FROM Requisition_Items ri
            JOIN Requisitions r ON r.id = ri.requisition_id
            WHERE r.lab_activity_date >= ? AND r.lab_activity_date <= ?
            GROUP BY ri.item_id
        """

        # Build result rows
        result = []
        for item in base_items:
            row = {
                "ITEMS": item["item_name"],
                "CATEGORIES": item["category"],
                "ACTUAL INVENTORY": item["actual_inventory"] or 0,
                "SIZE": item["size"] or "",
                "BRAND": item["brand"] or "",
                "OTHER SPECIFICATIONS": item["other_specifications"] or "",
            }

            item_grades = grade_usage_map.get(item["item_id"], {})
            row["GRADE 7"] = item_grades.get("Grade 7", 0)
            row["GRADE 8"] = item_grades.get("Grade 8", 0)
            row["GRADE 9"] = item_grades.get("Grade 9", 0)
            row["GRADE 10"] = item_grades.get("Grade 10", 0)
            row["TOTAL GRADE USAGE"] = (
                row["GRADE 7"] + row["GRADE 8"] + row["GRADE 9"] + row["GRADE 10"]
            )

            total_usage = 0
            for week in weeks:
                week_key = week.label  # "PRE", "WEEK 1", etc.

                # Query usage for this item in this week
                week_usage = (
                    db.execute_query(
                        usage_query, (week.start.isoformat(), week.end.isoformat())
                    )
                    or []
                )

                # Find this item's usage
                item_usage = 0
                for usage in week_usage:
                    if usage["item_id"] == item["item_id"]:
                        item_usage = usage["quantity"] or 0
                        break

                row[week_key] = item_usage
                total_usage += item_usage

            row["Total Number of Usage per Item"] = total_usage
            # Only include items with actual usage in the monthly report
            if total_usage > 0:
                result.append(row)

        return result, weeks

    except Exception as e:
        logger.error(f"Failed to get monthly usage data: {e}")
        return [], []


def create_monthly_usage_excel(
    data: List[Dict],
    weeks: List[WeekPeriod],
    output_path: Path,
    year: int,
    month: int,
    report_style: str = "detailed",
) -> None:
    """
    Create Excel file with monthly usage report in the exact beta test format.

    Layout:
    - Row 1: Empty (spacing)
    - Row 2: Header in columns F-G merged
    - Row 3: Empty (spacing)
    - Row 4: Column headers + month name over week columns
    - Row 5: "WEEK 1", "WEEK 2", etc. headers + "Total Number of Usage per Item"
    - Row 6: Date range labels under each week
    - Row 7+: Data rows with category headers

    Args:
        data: Report data rows
        weeks: Week definitions (WeekPeriod tuples)
        output_path: Output file path
        year: Report year
        month: Report month
        report_style: "detailed" for full title, "simple" for short title
    """
    try:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("Could not create worksheet")

        ws.title = "Monthly Usage"

        # Month names
        month_names_full = [
            "JANUARY",
            "FEBRUARY",
            "MARCH",
            "APRIL",
            "MAY",
            "JUNE",
            "JULY",
            "AUGUST",
            "SEPTEMBER",
            "OCTOBER",
            "NOVEMBER",
            "DECEMBER",
        ]
        month_name = month_names_full[month - 1]

        # Styles
        title_font = Font(bold=True, size=9)  # Reduced size to fit in cell
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(
            start_color="C65911", end_color="C65911", fill_type="solid"
        )  # Orange/brown header
        week_header_font = Font(bold=True, color="FFFFFF", size=9)
        date_range_font = Font(italic=True, size=9, color="008000")  # Green italic
        category_header_font = Font(bold=True, size=11)
        category_header_fill = PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        )  # Gold
        data_font = Font(size=10)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # Base column headers (A-F)
        base_headers = [
            "ITEMS",
            "CATEGORIES",
            "ACTUAL INVENTORY",
            "SIZE",
            "BRAND",
            "OTHER SPECIFICATIONS",
            "GRADE 7",
            "GRADE 8",
            "GRADE 9",
            "GRADE 10",
            "TOTAL GRADE USAGE",
        ]

        # Week headers from week definitions
        week_labels = [w.label for w in weeks]  # "PRE", "WEEK 1", etc.
        week_dates = [w.date_range for w in weeks]  # "Oct 3-7", etc.

        # === ROW 1: Empty ===
        # (No content needed)

        # === ROW 2: Title Header in F-G merged ===
        title_text = (
            "REPORT ON THE USAGE OF LABORATORY MATERIALS,\n"
            "EQUIPMENT AND APPRATUSES, ETC.\n"
            f"FOR THE MONTH OF {month_name} {year}"
        )
        ws["F2"] = title_text
        ws["F2"].font = title_font
        ws["F2"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.merge_cells("F2:G2")
        ws.row_dimensions[2].height = 50

        # === ROW 3: Empty ===
        # (No content needed)

        # === ROW 4-6: Base headers merged vertically + Month/Week headers ===
        week_start_col = len(base_headers) + 1
        week_end_col = week_start_col + len(weeks) - 1
        total_col = week_end_col + 1

        # Write base headers (A-F) merged across rows 4-6
        for col_num, header in enumerate(base_headers, 1):
            # Write value in row 4
            cell = ws.cell(row=4, column=col_num)
            cell.value = header  # type: ignore[assignment]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            # Merge rows 4-6 for this column
            ws.merge_cells(
                start_row=4, start_column=col_num, end_row=6, end_column=col_num
            )
            # Apply border to merged cells (rows 5-6)
            for row in range(5, 7):
                merge_cell = ws.cell(row=row, column=col_num)
                merge_cell.border = border

        # Month name header over week columns (row 4, columns G onwards)
        month_cell = ws.cell(row=4, column=week_start_col)
        month_cell.value = f"{month_name.lower()} {year}"  # type: ignore[assignment]
        month_cell.font = header_font  # Same style as other headers
        month_cell.fill = header_fill  # Same orange color
        month_cell.alignment = center_align
        month_cell.border = border

        if len(weeks) > 1:
            ws.merge_cells(
                start_row=4,
                start_column=week_start_col,
                end_row=4,
                end_column=week_end_col,
            )
            # Apply border to merged month cells
            for col in range(week_start_col, week_end_col + 1):
                ws.cell(row=4, column=col).border = border

        # "Total Number of Usage per Item" header merged rows 4-6
        total_cell = ws.cell(row=4, column=total_col)
        total_cell.value = "Total Number of\nUsage per Item"  # type: ignore[assignment]
        total_cell.font = header_font
        total_cell.fill = header_fill
        total_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        total_cell.border = border
        # Merge rows 4-6 for total column
        ws.merge_cells(
            start_row=4, start_column=total_col, end_row=6, end_column=total_col
        )
        # Apply border to merged cells
        for row in range(5, 7):
            ws.cell(row=row, column=total_col).border = border

        # === ROW 5: Week labels (PRE, WEEK 1, etc.) ===
        for i, week_label in enumerate(week_labels):
            col = week_start_col + i
            cell = ws.cell(row=5, column=col)
            cell.value = week_label  # type: ignore[assignment]
            cell.font = week_header_font
            cell.fill = header_fill  # Same orange color
            cell.alignment = center_align
            cell.border = border

        # === ROW 6: Date ranges under each week ===
        for i, week_date in enumerate(week_dates):
            col = week_start_col + i
            cell = ws.cell(row=6, column=col)
            cell.value = week_date  # type: ignore[assignment]
            cell.font = date_range_font
            cell.alignment = center_align
            cell.border = border

        # Set column widths
        col_widths = {
            1: 25,  # ITEMS
            2: 15,  # CATEGORIES
            3: 18,  # ACTUAL INVENTORY
            4: 12,  # SIZE
            5: 15,  # BRAND
            6: 20,  # OTHER SPECIFICATIONS
            7: 10,  # GRADE 7
            8: 10,  # GRADE 8
            9: 10,  # GRADE 9
            10: 10,  # GRADE 10
            11: 16,  # TOTAL GRADE USAGE
        }
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Week columns width
        for i in range(len(weeks)):
            ws.column_dimensions[get_column_letter(week_start_col + i)].width = 12

        # Total column width
        ws.column_dimensions[get_column_letter(total_col)].width = 18

        # Group data by category
        grouped_data: Dict[str, List[Dict]] = OrderedDict()
        for cat in CATEGORY_ORDER:
            grouped_data[cat] = []

        for row in data:
            category = row.get("CATEGORIES", "Uncategorized")
            if category not in grouped_data:
                grouped_data[category] = []
            grouped_data[category].append(row)

        # Build all headers for data writing
        all_headers = base_headers + week_labels + ["Total Number of Usage per Item"]

        # Write data rows starting at row 7
        current_row = 7

        for category in CATEGORY_ORDER:
            items = grouped_data.get(category, [])
            if not items:
                continue

            # Category header row
            display_name = CATEGORY_DISPLAY_NAMES.get(category, category.upper())
            cell = ws.cell(row=current_row, column=1)
            cell.value = display_name  # type: ignore[assignment]
            cell.font = category_header_font
            cell.fill = category_header_fill
            cell.border = border

            # Fill rest of category header row
            for col in range(2, len(all_headers) + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = category_header_fill
                cell.border = border

            current_row += 1

            # Data rows for this category
            for item in items:
                for col_num, header in enumerate(all_headers, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    value = item.get(header, "")
                    cell.value = value if value is not None else ""  # type: ignore[assignment]
                    cell.font = data_font
                    cell.border = border

                    # Align numbers to center, text to left
                    if col_num > len(base_headers):
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align

                current_row += 1

        # Freeze panes at row 7 (data starts here)
        ws.freeze_panes = ws.cell(row=7, column=1)  # type: ignore[assignment]

        # Save workbook
        wb.save(output_path)
        logger.info(f"Monthly usage report saved to {output_path}")

    except Exception as e:
        logger.error(f"Failed to create monthly usage Excel report: {e}")
        raise


def generate_monthly_usage_report(
    year: int,
    month: int,
    output_path: Optional[str] = None,
    category_filter: str = "",
    report_style: str = "detailed",
) -> str:
    """
    Generate monthly usage report in Excel format.

    Args:
        year: Report year
        month: Report month (1-12)
        output_path: Optional output file path
        category_filter: Optional category filter
        report_style: "detailed" or "simple" title style

    Returns:
        Path to generated Excel file or error message
    """
    try:
        logger.info(f"Generating monthly usage report for {year}-{month:02d}")

        # Get data
        data, weeks = get_monthly_usage_data(year, month, category_filter)

        if not data:
            return f"No data found for {year}-{month:02d}"

        # Generate output path if not provided
        if not output_path:
            month_names = [
                "jan",
                "feb",
                "mar",
                "apr",
                "may",
                "jun",
                "jul",
                "aug",
                "sep",
                "oct",
                "nov",
                "dec",
            ]
            month_name = month_names[month - 1]
            from datetime import datetime

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"monthly_usage_{month_name}_{year}_{timestamp}.xlsx"

        output_path_obj = Path(output_path)

        # Create Excel report
        create_monthly_usage_excel(
            data, weeks, output_path_obj, year, month, report_style
        )

        return str(output_path)

    except Exception as e:
        logger.error(f"Failed to generate monthly usage report: {e}")
        return f"Error generating report: {e}"

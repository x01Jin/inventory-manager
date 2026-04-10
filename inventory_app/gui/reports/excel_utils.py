from typing import List, Dict, Optional, cast
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell

from inventory_app.gui.reports.header_utils import format_excel_headers
from inventory_app.utils.logger import logger
from inventory_app.gui.reports.report_utils import date_formatter


# Category display order matching sample reports
CATEGORY_ORDER = [
    "Equipment",
    "Apparatus",
    "Lab Models",
    "Chemicals-Solid",
    "Chemicals-Liquid",
    "Prepared Slides",
    "Consumables",
    "Others",
    "Uncategorized",
]

# Category display names for report headers
CATEGORY_DISPLAY_NAMES = {
    "Equipment": "EQUIPMENT",
    "Apparatus": "APPARATUSES",
    "Lab Models": "LAB MODELS",
    "Chemicals-Solid": "CHEMICALS - SOLID",
    "Chemicals-Liquid": "CHEMICALS - LIQUID",
    "Prepared Slides": "PREPARED SLIDES",
    "Consumables": "CONSUMABLES",
    "Others": "OTHERS",
    "Uncategorized": "UNCATEGORIZED",
}


def create_excel_report(
    data: List[Dict],
    output_path: Path,
    title: str,
    start_date: date,
    end_date: date,
    granularity: Optional[str] = None,
) -> None:
    """
    Create Excel file with report data in the standardized styled format.
    """
    try:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("Could not create worksheet")

        ws.title = "Report"

        # Styles
        title_font = Font(bold=True, size=11)
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(
            start_color="C65911", end_color="C65911", fill_type="solid"
        )  # Orange/brown header

        category_header_font = Font(bold=True, size=11)
        category_header_fill = PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        )  # Gold
        data_font = Font(size=10)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        if not data:
            ws["A1"] = "No data found for the specified period."
            wb.save(output_path)
            return

        # Prepare headers
        raw_headers = list(data[0].keys())
        formatted_headers = format_excel_headers(
            raw_headers, start_date, end_date, granularity
        )

        # === ROW 1: Empty ===

        # === ROW 2: Title Header ===
        period_desc = date_formatter.get_date_range_description(start_date, end_date)
        full_title = f"{title.upper()}\nPERIOD: {period_desc.upper()}"

        # Determine title merge range (centered over first few columns)
        merge_end_col = min(7, len(formatted_headers))
        title_cell = cast(Cell, ws.cell(row=2, column=1))
        title_cell.value = full_title
        title_cell.font = title_font
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.row_dimensions[2].height = 40
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_end_col)

        # === ROW 3: Empty ===

        # === ROW 4: Headers ===
        for col_num, header_text in enumerate(formatted_headers, 1):
            cell = cast(Cell, ws.cell(row=4, column=col_num))
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

            # Set column width
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(
                len(str(header_text)) + 4, 15
            )

        # Auto-filter across headers and data
        try:
            last_col_letter = get_column_letter(len(formatted_headers))
            ws.auto_filter.ref = (
                f"A4:{last_col_letter}{4 + len(data) + 10}"  # +10 for categories
            )
        except Exception:
            pass

        # Group data by category if CATEGORIES column exists
        category_col_key = None
        for key in ["CATEGORIES", "Category", "category"]:
            if key in data[0]:
                category_col_key = key
                break

        current_row = 5
        if category_col_key:
            from collections import defaultdict

            grouped_data = defaultdict(list)
            for row in data:
                cat = row.get(category_col_key, "Uncategorized")
                grouped_data[cat].append(row)

            # Sort categories by defined order
            sorted_categories = []
            for cat in CATEGORY_ORDER:
                if cat in grouped_data:
                    sorted_categories.append(cat)

            # Add any other categories not in CATEGORY_ORDER
            for cat in grouped_data.keys():
                if cat not in CATEGORY_ORDER:
                    sorted_categories.append(cat)

            for category in sorted_categories:
                items = grouped_data[category]

                # Category Header Row
                display_name = CATEGORY_DISPLAY_NAMES.get(
                    category, str(category).upper()
                )
                cat_cell = cast(Cell, ws.cell(row=current_row, column=1))
                cat_cell.value = display_name
                cat_cell.font = category_header_font
                cat_cell.fill = category_header_fill
                cat_cell.border = border

                for col in range(2, len(formatted_headers) + 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = category_header_fill
                    cell.border = border

                ws.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=len(formatted_headers),
                )
                current_row += 1

                # Data Rows
                for row_data in items:
                    for col_num, value in enumerate(row_data.values(), 1):
                        cell = cast(Cell, ws.cell(row=current_row, column=col_num))
                        cell.value = value
                        cell.font = data_font
                        cell.border = border

                        # Right-align numbers and apply formatting
                        if isinstance(value, (int, float)):
                            cell.alignment = Alignment(horizontal="right")
                            cell.number_format = "#,##0"
                        else:
                            cell.alignment = left_align
                    current_row += 1
        else:
            # Just write data sequentially if no category grouping
            for row_data in data:
                for col_num, value in enumerate(row_data.values(), 1):
                    cell = cast(Cell, ws.cell(row=current_row, column=col_num))
                    cell.value = value
                    cell.font = data_font
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = "#,##0"
                    else:
                        cell.alignment = left_align
                current_row += 1

        # Add grand total row
        try:
            # Determine numeric columns
            numeric_indices = set()
            for idx, head in enumerate(formatted_headers, 1):
                low = str(head).lower() if head else ""
                if any(
                    x in low for x in ["quantity", "stock", "total"]
                ) or low.endswith("qty"):
                    numeric_indices.add(idx)

            if numeric_indices:
                total_row = current_row
                total_font = Font(bold=True)

                label_cell = cast(Cell, ws.cell(row=total_row, column=1))
                label_cell.value = "Total"
                label_cell.font = total_font
                label_cell.border = border

                # Apply borders to the whole total row
                for col in range(2, len(formatted_headers) + 1):
                    ws.cell(row=total_row, column=col).border = border

                # Sum columns
                for col_idx in numeric_indices:
                    header_key = raw_headers[col_idx - 1]
                    total_val = 0
                    for item in data:
                        try:
                            val = item.get(header_key, 0) or 0
                            total_val += float(val)
                        except (ValueError, TypeError):
                            pass

                    cell = cast(Cell, ws.cell(row=total_row, column=col_idx))
                    cell.value = total_val
                    cell.font = total_font
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                    cell.border = border

                current_row += 1
        except Exception as e:
            logger.debug(f"Failed to add totals row: {e}")

        # Freeze panes at row 5
        ws.freeze_panes = "A5"

        # Final column width adjustment
        for col_num in range(1, len(formatted_headers) + 1):
            column_letter = get_column_letter(col_num)
            max_len = ws.column_dimensions[column_letter].width
            # Scan a few rows for content length
            for r in range(5, min(current_row, 25)):
                val = ws.cell(row=r, column=col_num).value
                if val:
                    max_len = max(max_len, len(str(val)) + 2)
            ws.column_dimensions[column_letter].width = min(max_len, 50)

        wb.save(output_path)
        logger.info(f"Report saved to {output_path}")

    except Exception as e:
        logger.error(f"Failed to create Excel report: {e}")
        raise

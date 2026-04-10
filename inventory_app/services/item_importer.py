"""
Item importer service

Single-purpose module to import items from Excel (.xlsx) files. Validates headers and
normalizes values before creating Item records via the existing model API.

Supports progress callbacks for async UI updates.
"""

from typing import Dict, List, Set, Tuple, Any, Optional, Callable
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from inventory_app.database.models import Item
from inventory_app.utils.logger import logger
from inventory_app.database.connection import db
from datetime import datetime, date
import re


REQUIRED_HEADERS = {"name", "stocks", "stock", "item type", "item_type", "type"}

# Normalized header key map (space/underscore-insensitive, case-insensitive)


def _normalize_key_for_map(s: Any) -> str:
    s_str = str(s) if s is not None else ""
    # Keep only alphanumeric characters so 'item type', 'item_type' and
    # 'ItemType' all normalize to 'itemtype'.
    return "".join(ch for ch in s_str.lower() if ch.isalnum())


# Mapping of normalized header -> canonical field name used in code
HEADER_MAP = {
    "name": "name",
    "item": "name",
    "items": "name",
    "itemname": "name",
    "itemnames": "name",
    "names": "name",
    "stocks": "stocks",
    "stock": "stocks",
    "itemtype": "item_type",
    "type": "item_type",
    "category": "category",
    "size": "size",
    "brand": "brand",
    "supplier": "supplier",
    "otherspecifications": "other_specifications",
    "ponumber": "po_number",
    "expirationdate": "expiration_date",
    "calibrationdate": "calibration_date",
    "acquisitiondate": "acquisition_date",
}

# Build normalized header map once now that HEADER_MAP exists
HEADER_MAP_NORMALIZED = {_normalize_key_for_map(k): v for k, v in HEADER_MAP.items()}

# Acceptable name header variants (human-friendly forms). Normalized for comparisons below.
NAME_VARIANTS = [
    "name",
    "item",
    "items",
    "item name",
    "item names",
    "itemname",
    "itemnames",
    "names",
]
NAME_NORMALIZED = {_normalize_key_for_map(v) for v in NAME_VARIANTS}


def _normalize_header(h: Any) -> str:
    """Normalize a header value to a lowercase, space/underscore-insensitive string.

    Converts to str, lowercases, and removes spaces and underscores. Returns
    an empty string for missing/invalid values.
    """
    if h is None:
        return ""
    s = str(h).strip().lower()
    if s == "":
        return ""
    # Remove spaces and underscores to make header matching space/underscore-insensitive
    return s.replace(" ", "").replace("_", "")


def _normalize_item_type(raw_item_type: Any) -> str:
    """Normalize item type text used for consumable classification."""
    raw_type_text = "" if raw_item_type is None else str(raw_item_type)
    return re.sub(r"^\s*ta,\s*", "", raw_type_text, flags=re.I).strip().lower()


def _to_canonical_item_type(raw_item_type: Any) -> str:
    """Map raw item type text to canonical stored values."""
    raw_text = "" if raw_item_type is None else str(raw_item_type)
    stripped = raw_text.strip()
    lowered = stripped.lower()

    if (
        "ta" in lowered
        and ("non" in lowered or "not" in lowered)
        and "consum" in lowered
    ):
        return "TA, non-consumable"

    if _is_consumable_type(raw_item_type):
        return "Consumable"

    if ("non" in lowered or "not" in lowered) and "consum" in lowered:
        return "Non-consumable"

    return stripped or "Non-consumable"


def _resolve_supplier_id(raw_supplier: Any) -> Optional[int]:
    """Resolve supplier text into Suppliers.id, creating a record when missing."""
    if raw_supplier is None:
        return None

    supplier_name = str(raw_supplier).strip()
    if supplier_name == "" or supplier_name.upper() == "N/A":
        return None

    try:
        existing = db.execute_query(
            "SELECT id FROM Suppliers WHERE name = ? COLLATE NOCASE",
            (supplier_name,),
        )
        if existing:
            return int(existing[0]["id"])

        db.execute_update("INSERT INTO Suppliers (name) VALUES (?)", (supplier_name,))
        created = db.execute_query(
            "SELECT id FROM Suppliers WHERE name = ? COLLATE NOCASE",
            (supplier_name,),
        )
        if created:
            return int(created[0]["id"])
    except Exception as exc:
        logger.warning(f"Could not resolve supplier '{supplier_name}': {exc}")

    return None


def _is_consumable_type(raw_item_type: Any) -> bool:
    """Return True when raw item type text resolves to consumable."""
    cleaned = _normalize_item_type(raw_item_type)
    if ("non" in cleaned or "not" in cleaned) and "consum" in cleaned:
        return False

    return any(
        k in cleaned
        for k in (
            "consum",
            "consumable",
            "consumables",
            "reagent",
            "reagents",
            "chemical",
        )
    )


def _resolve_header_mapping(ws: Worksheet) -> Tuple[int, Dict[int, str], List[str]]:
    """Find header row and build a column index -> canonical field mapping."""
    max_scan_rows = min(40, ws.max_row if ws.max_row is not None else 40)
    header_row_index = None
    headers: List[str] = []

    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        if not row:
            continue

        scanned = [_normalize_header(cell) for cell in row]
        normalized_header_set = {h for h in scanned if h}

        has_name = bool(NAME_NORMALIZED & normalized_header_set)
        has_stocks = bool({"stocks", "stock"} & normalized_header_set)
        has_item_type = bool({"itemtype", "type"} & normalized_header_set)

        if has_name and has_stocks and has_item_type:
            headers = scanned
            header_row_index = i
            break

    if header_row_index is None:
        raise ValueError(
            "Missing required header row (need columns for name, stocks, and item type). Check if all required headers are present and double check the spelling of the headers"
        )

    col_map: Dict[int, str] = {}
    for idx, h in enumerate(headers, start=1):
        if not h:
            continue
        mapped = HEADER_MAP_NORMALIZED.get(h)
        if mapped:
            col_map[idx] = mapped

    normalized_header_set = {h for h in headers if h}
    if not (NAME_NORMALIZED & normalized_header_set):
        raise ValueError(
            "Missing required header 'name' (or acceptable variants like 'item' or 'item name'). Check if all required headers are present and double check the spelling of the headers"
        )
    if not ({"stocks", "stock"} & normalized_header_set):
        raise ValueError(
            "Missing required header 'stocks'. Check if all required headers are present and double check the spelling of the headers"
        )
    if not ({"itemtype", "type"} & normalized_header_set):
        raise ValueError(
            "Missing required header 'item type'. Check if all required headers are present and double check the spelling of the headers"
        )

    return header_row_index, col_map, headers


def _is_decimal_without_unit(raw_stocks: Any) -> bool:
    """Return True for numeric decimal stocks with no explicit unit text."""
    if isinstance(raw_stocks, float):
        return not float(raw_stocks).is_integer()
    if isinstance(raw_stocks, int):
        return False

    if raw_stocks is None:
        return False

    text = str(raw_stocks).strip()
    if text == "":
        return False

    # Numeric-only strings like "1.5" are ambiguous for consumables.
    if re.fullmatch(r"\d+\.\d+", text):
        return True

    return False


def _with_selected_unit(raw_stocks: Any, unit: str) -> Any:
    """Attach a selected unit to raw stock values before parsing."""
    if raw_stocks is None:
        return raw_stocks
    if isinstance(raw_stocks, (int, float)):
        return f"{raw_stocks} {unit}"

    text = str(raw_stocks).strip()
    if text == "":
        return raw_stocks
    return f"{text} {unit}"


def collect_consumable_rows_missing_unit(path: str) -> List[Dict[str, Any]]:
    """List consumable rows with decimal stock values that have no explicit unit.

    Returns list entries with keys: row_index, name, stocks.
    """
    wb = load_workbook(path, data_only=True)
    ws_any = wb.active
    if ws_any is None:
        raise ValueError("The Excel workbook has no active worksheet")

    ws: Worksheet = ws_any  # type: ignore[assignment]
    header_row_index, col_map, _ = _resolve_header_mapping(ws)

    flagged_rows: List[Dict[str, Any]] = []
    start_row = int(header_row_index) + 1

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=start_row, values_only=True), start=start_row
    ):
        row_data: Dict[str, Any] = {}
        for col_idx, cell_value in enumerate(row, start=1):
            if col_idx in col_map:
                row_data[col_map[col_idx]] = cell_value

        name = row_data.get("name")
        item_type_raw = row_data.get("item_type")
        raw_stocks = row_data.get("stocks")

        if name is None or str(name).strip() == "":
            continue
        if not _is_consumable_type(item_type_raw):
            continue
        if not _is_decimal_without_unit(raw_stocks):
            continue

        flagged_rows.append(
            {
                "row_index": row_idx,
                "name": str(name).strip(),
                "stocks": str(raw_stocks).strip(),
            }
        )

    return flagged_rows


def _parse_int(val) -> int:
    try:
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return 0
        return int(float(val))
    except Exception:
        raise ValueError(f"Invalid stock value: {val}")


def _parse_date(val) -> Optional[date]:
    """Return a date object or None.

    Accepts datetime/date objects or parseable date strings. Returns None for
    empty or unparseable values.
    """
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # openpyxl may return date/datetime objects or strings
    try:
        dt = datetime.fromisoformat(str(val))
        return dt.date()
    except Exception:
        # if not parseable, return None and let caller handle
        return None


def import_items_from_excel(
    path: str,
    editor_name: str = "Import",
    progress_callback: Optional[Callable[[int, int, int], None]] = None,
    row_unit_overrides: Optional[Dict[int, str]] = None,
    rows_to_skip: Optional[Set[int]] = None,
) -> Tuple[int, List[str]]:
    """Import items from an Excel file.

    Args:
        path: Path to .xlsx file
        editor_name: Name stored in audit logs
        progress_callback: Optional callback(current, total, skipped) for progress updates
        row_unit_overrides: Optional mapping of Excel row index -> selected unit
            (e.g., {42: "L"}) for consumables with missing units.
        rows_to_skip: Optional set of Excel row indexes to skip by user choice.

    Returns:
        (count_imported, messages): number of successfully imported rows and list of messages

    Raises ValueError with human-friendly message when headers are missing.
    """
    wb = load_workbook(path, data_only=True)
    ws_any = wb.active
    if ws_any is None:
        raise ValueError("The Excel workbook has no active worksheet")

    # Explicitly assert typing for linters; active() may return a types union
    ws: Worksheet = ws_any  # type: ignore[assignment]

    header_row_index, col_map, headers = _resolve_header_mapping(ws)
    logger.debug(f"Header found on row {header_row_index}: {headers}")
    logger.debug(f"Column mapping: {col_map}")

    if row_unit_overrides is None:
        row_unit_overrides = {}
    if rows_to_skip is None:
        rows_to_skip = set()

    messages: List[str] = []
    imported = 0
    skipped = 0

    try:
        # Count total rows to process for progress reporting
        assert header_row_index is not None
        start_row = int(header_row_index) + 1
        total_rows = (ws.max_row or start_row) - start_row + 1
        if total_rows < 0:
            total_rows = 0

        current_row = 0

        # No global transaction: save() handles its own transaction. This prevents
        # nested transaction errors when the DB driver does not support them.
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=start_row, values_only=True), start=start_row
        ):
            current_row += 1

            # Report progress
            if progress_callback:
                progress_callback(current_row, total_rows, skipped)
            # Build a row dict using mapping (values_only=True returns raw values)
            row_data = {}
            for col_idx, cell_value in enumerate(row, start=1):
                if col_idx in col_map:
                    key = col_map[col_idx]
                    row_data[key] = cell_value

            # Required fields
            name = row_data.get("name")
            raw_stocks = row_data.get("stocks")
            item_type_raw = row_data.get("item_type")

            if name is None or str(name).strip() == "":
                messages.append(
                    f"Row {row_idx}: missing required 'name' value; skipping"
                )
                skipped += 1
                continue

            if row_idx in rows_to_skip:
                messages.append(
                    f"Row {row_idx}: skipped by user during unit resolution"
                )
                skipped += 1
                continue

            selected_unit = row_unit_overrides.get(row_idx)
            if selected_unit:
                raw_stocks = _with_selected_unit(raw_stocks, selected_unit)

            # Parse the free-form stocks cell. This may extract quantity, a size string
            # (e.g. '900ml', '1 L') and an optional notes string. For size-bearing
            # entries (volume/mass), quantity comes from the numeric part so consumables
            # can be requested/returned in partial usable amounts.
            size_from_stocks = None
            try:
                from inventory_app.utils.stock_parser import parse_stock_value

                stock_info = parse_stock_value(raw_stocks)
                stocks = int(stock_info.get("quantity", 0))
                size_from_stocks = stock_info.get("size")
                stocks_notes = stock_info.get("notes")
            except ValueError as e:
                messages.append(f"Row {row_idx}: {str(e)}; skipping")
                skipped += 1
                continue

            if item_type_raw is None or str(item_type_raw).strip() == "":
                messages.append(
                    f"Row {row_idx}: missing required 'item type' value; skipping"
                )
                skipped += 1
                continue

            # Normalize and clean item type text.
            # Some vendor spreadsheets include a leading code like 'TA, ' before the
            # actual item type (e.g., 'TA, non consumable'). Strip such prefixes
            # case-insensitively and robustly detect consumable vs non-consumable.
            canonical_item_type = _to_canonical_item_type(item_type_raw)
            is_consumable = 1 if _is_consumable_type(item_type_raw) else 0

            # Optional fields (text fields default to 'N/A'; date fields return date objects or None)
            size = row_data.get("size")
            brand = row_data.get("brand") or "N/A"
            other_specifications = row_data.get("other_specifications") or "N/A"
            po_number = row_data.get("po_number") or "N/A"

            # If the 'stocks' cell contained a size (e.g. '900ml') and the explicit
            # size column is empty, prefer the size parsed from stocks.
            if (size is None or str(size).strip() == "") and size_from_stocks:
                size = size_from_stocks

            # If the stock parser returned notes (e.g., '(100pcs)' or 'of 8 pieces'),
            # append them to other_specifications for visibility.
            if stocks_notes:
                # Avoid duplicating 'N/A' as the literal value
                if other_specifications in (None, "", "N/A"):
                    other_specifications = str(stocks_notes)
                else:
                    other_specifications = f"{other_specifications}; {stocks_notes}"

            # Ensure size gets a benign default for the existing save() semantics
            if size is None or str(size).strip() == "":
                size = "N/A"

            expiration_date = _parse_date(row_data.get("expiration_date"))
            calibration_date = _parse_date(row_data.get("calibration_date"))
            acquisition_date = _parse_date(row_data.get("acquisition_date"))
            supplier_id = _resolve_supplier_id(row_data.get("supplier"))

            # Resolve category: try to match existing Category by name (case-insensitive).
            # Categories are fixed and read-only (as of v0.7.0b); unknown categories are
            # mapped to 'Uncategorized'. If no category provided or value is empty/N/A,
            # also use 'Uncategorized'.
            raw_category = row_data.get("category")
            try:
                category_id = None
                if raw_category and str(raw_category).strip().upper() not in (
                    "",
                    "N/A",
                ):
                    cat_name = str(raw_category).strip()
                    rows = db.execute_query(
                        "SELECT id FROM Categories WHERE name = ? COLLATE NOCASE",
                        (cat_name,),
                    )
                    if rows:
                        category_id = rows[0]["id"]
                    else:
                        # Category doesn't exist; log and map to Uncategorized
                        logger.debug(
                            f"Category '{cat_name}' not found; mapping to 'Uncategorized'"
                        )

                # If no category was matched or if the provided category doesn't exist,
                # use 'Uncategorized'
                if category_id is None:
                    rows = db.execute_query(
                        "SELECT id FROM Categories WHERE name = ? COLLATE NOCASE",
                        ("Uncategorized",),
                    )
                    if rows:
                        category_id = rows[0]["id"]
                    else:
                        # Fallback to first available category if Uncategorized doesn't exist
                        fallback = db.execute_query(
                            "SELECT id FROM Categories ORDER BY id LIMIT 1"
                        )
                        if fallback:
                            category_id = fallback[0]["id"]
            except Exception as e:
                logger.error(f"Failed to resolve category '{raw_category}': {e}")
                # As a final fallback, use first category id if available
                fallback = db.execute_query(
                    "SELECT id FROM Categories ORDER BY id LIMIT 1"
                )
                category_id = fallback[0]["id"] if fallback else 1

            # Ensure category_id is an int (type guard for static analysis and DB)
            if category_id is None:
                category_id = 1
            elif not isinstance(category_id, int):
                try:
                    category_id = int(category_id)
                except Exception:
                    logger.warning(
                        f"Category id {category_id!r} could not be coerced to int; using 1 as fallback"
                    )
                    category_id = 1

            # Create and save the item
            item = Item(
                name=str(name).strip(),
                category_id=category_id,
                item_type=canonical_item_type,
                size=size if size != "" else None,
                brand=brand if brand != "" else None,
                other_specifications=other_specifications
                if other_specifications != ""
                else None,
                po_number=po_number if po_number != "" else None,
                supplier_id=supplier_id,
                expiration_date=expiration_date,
                calibration_date=calibration_date,
                is_consumable=is_consumable,
                acquisition_date=acquisition_date,
            )

            # Log the parameters we will try to save (helps diagnose FK/constraint failures)
            try:
                save_params = (
                    item.name,
                    item.category_id,
                    item.size,
                    item.brand,
                    item.other_specifications,
                    item.po_number,
                    item.supplier_id,
                    item.item_type,
                    item.expiration_date,
                    item.calibration_date,
                    item.is_consumable,
                    item.acquisition_date,
                )
                logger.debug(f"Saving item (row {row_idx}) params: {save_params}")
            except Exception:
                logger.debug(
                    "Saving item (row {row_idx}) - unable to serialize params for logging"
                )

            try:
                # item.save() will guard and return False on failures; still protect
                # against unexpected exceptions to resume importing other rows.
                success = item.save(editor_name=editor_name, batch_quantity=stocks)
                if success:
                    imported += 1
                    # include parsed size in messages when available
                    msg_size = f", size={size}" if size and size != "N/A" else ""
                    messages.append(
                        f"Row {row_idx}: imported '{item.name}' ({stocks} units{msg_size})"
                    )
                else:
                    skipped += 1
                    messages.append(
                        f"Row {row_idx}: failed to import '{item.name}' (db error)"
                    )
            except Exception as e:
                logger.error(f"Exception saving item on row {row_idx}: {e}")
                skipped += 1
                messages.append(
                    f"Row {row_idx}: exception during save: {str(e)}; skipping"
                )

        logger.info(f"Import completed: {imported} items added from {path}")
        return imported, messages
    except Exception as ex:
        logger.error(f"Import failed: {ex}")
        raise

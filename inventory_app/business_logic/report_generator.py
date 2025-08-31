"""
Report generator for the inventory application.
Handles Excel report generation for weekly, monthly, quarterly, and yearly reports.
Uses composition pattern with DatabaseConnection and openpyxl.
"""

from typing import List, Dict, Optional
from datetime import date, datetime
from pathlib import Path
import calendar

from inventory_app.database.connection import db
from inventory_app.utils.logger import logger

# Import openpyxl directly since it's in requirements.txt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """
    Handles report generation and Excel export.
    Uses composition with DatabaseConnection and openpyxl.
    """

    def __init__(self):
        """Initialize report generator."""
        logger.info("Report generator initialized")

    def generate_weekly_report(self, start_date: date, end_date: date,
                              output_path: Optional[str] = None) -> str:
        """
        Generate weekly usage report in Excel format.

        Args:
            start_date: Start date for the report period
            end_date: End date for the report period
            output_path: Optional output file path

        Returns:
            Path to generated Excel file
        """
        try:
            logger.info(f"Generating weekly report from {start_date} to {end_date}")

            # Get report data
            report_data = self._get_weekly_report_data(start_date, end_date)

            if not report_data:
                logger.warning("No data found for the specified period")
                return ""

            # Create Excel file
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"weekly_report_{timestamp}.xlsx"

            output_path_obj = Path(output_path)
            self._create_excel_report(report_data, output_path_obj, "Weekly Usage Report",
                                    start_date, end_date)

            logger.info(f"Weekly report generated: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Failed to generate weekly report: {e}")
            return ""

    def generate_monthly_report(self, year: int, month: int,
                               output_path: Optional[str] = None) -> str:
        """
        Generate monthly usage report.

        Args:
            year: Year for the report
            month: Month for the report (1-12)
            output_path: Optional output file path

        Returns:
            Path to generated Excel file
        """
        try:
            # Calculate start and end dates for the month
            start_date = date(year, month, 1)
            _, last_day = calendar.monthrange(year, month)
            end_date = date(year, month, last_day)

            return self.generate_weekly_report(start_date, end_date, output_path)

        except Exception as e:
            logger.error(f"Failed to generate monthly report: {e}")
            return ""

    def generate_quarterly_report(self, year: int, quarter: int,
                                 output_path: Optional[str] = None) -> str:
        """
        Generate quarterly usage report.

        Args:
            year: Year for the report
            quarter: Quarter (1-4)
            output_path: Optional output file path

        Returns:
            Path to generated Excel file
        """
        try:
            # Calculate start and end dates for the quarter
            quarter_start_months = {1: 1, 2: 4, 3: 7, 4: 10}
            start_month = quarter_start_months[quarter]
            end_month = start_month + 2

            start_date = date(year, start_month, 1)
            if end_month == 12:
                end_date = date(year, 12, 31)
            else:
                _, last_day = calendar.monthrange(year, end_month)
                end_date = date(year, end_month, last_day)

            return self.generate_weekly_report(start_date, end_date, output_path)

        except Exception as e:
            logger.error(f"Failed to generate quarterly report: {e}")
            return ""

    def generate_yearly_report(self, year: int, output_path: Optional[str] = None) -> str:
        """
        Generate yearly usage report.

        Args:
            year: Year for the report
            output_path: Optional output file path

        Returns:
            Path to generated Excel file
        """
        try:
            start_date = date(year, 1, 1)
            end_date = date(year, 12, 31)

            return self.generate_weekly_report(start_date, end_date, output_path)

        except Exception as e:
            logger.error(f"Failed to generate yearly report: {e}")
            return ""

    def _get_weekly_report_data(self, start_date: date, end_date: date) -> List[Dict]:
        """
        Get weekly report data from database.

        Args:
            start_date: Start date for the report
            end_date: End date for the report

        Returns:
            List of report rows
        """
        try:
            # Use the weekly report query from schema.sql
            query = """
            WITH base AS (
              SELECT
                i.id AS item_id,
                i.name AS item_name,
                c.name AS category_name,
                i.size,
                i.brand,
                i.other_specifications,
                SUM(ri.quantity_borrowed) AS qty,
                r.lab_activity_date
              FROM Requisition_Items ri
              JOIN Requisitions r ON r.id = ri.requisition_id
              JOIN Items i ON i.id = ri.item_id
              JOIN Categories c ON c.id = i.category_id
              WHERE r.lab_activity_date BETWEEN ? AND ?
              GROUP BY i.id, r.lab_activity_date
            ),
            with_weeks AS (
              SELECT
                item_id, item_name, category_name, size, brand, other_specifications, qty, lab_activity_date,
                CAST((strftime('%d', lab_activity_date) - 1) / 7 + 1 AS INTEGER) AS week_of_month
              FROM base
            )
            SELECT
              item_name AS ITEMS,
              category_name AS CATEGORIES,
              (SELECT COALESCE(SUM(quantity_received), 0)
                 FROM Item_Batches b
                 WHERE b.item_id = w.item_id) AS ACTUAL_INVENTORY,
              size AS SIZE,
              brand AS BRAND,
              other_specifications AS "OTHER SPECIFICATIONS",
              SUM(CASE WHEN week_of_month = 1 THEN qty ELSE 0 END) AS "WEEK 1",
              SUM(CASE WHEN week_of_month = 2 THEN qty ELSE 0 END) AS "WEEK 2",
              SUM(CASE WHEN week_of_month = 3 THEN qty ELSE 0 END) AS "WEEK 3",
              SUM(CASE WHEN week_of_month = 4 THEN qty ELSE 0 END) AS "WEEK 4",
              SUM(CASE WHEN week_of_month = 5 THEN qty ELSE 0 END) AS "WEEK 5",
              SUM(qty) AS "TOTAL NUMBER OF ITEMS"
            FROM with_weeks w
            GROUP BY item_id
            ORDER BY category_name, item_name
            """

            rows = db.execute_query(query, (start_date.isoformat(), end_date.isoformat()))
            return rows

        except Exception as e:
            logger.error(f"Failed to get weekly report data: {e}")
            return []

    def _create_excel_report(self, data: List[Dict], output_path: Path,
                           title: str, start_date: date, end_date: date) -> None:
        """
        Create Excel file with report data.

        Args:
            data: Report data rows
            output_path: Output file path
            title: Report title
            start_date: Report start date
            end_date: Report end date
        """
        try:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("Could not create worksheet")

            ws.title = "Report"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center')

            # Add title
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            ws['A2'].font = Font(italic=True)

            # Add headers
            if data:
                headers = list(data[0].keys())
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col_num)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                    # Auto-adjust column width
                    column_letter = get_column_letter(col_num)
                    ws.column_dimensions[column_letter].width = max(len(str(header)) + 2, 12)

            # Add data rows
            for row_num, row_data in enumerate(data, 5):
                for col_num, value in enumerate(row_data.values(), 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border

                    # Auto-adjust column width based on content
                    column_letter = get_column_letter(col_num)
                    content_length = len(str(value)) + 2
                    current_width = ws.column_dimensions[column_letter].width or 12
                    ws.column_dimensions[column_letter].width = max(current_width, content_length)

            # Save the workbook
            wb.save(output_path)
            logger.debug(f"Excel file saved to {output_path}")

        except Exception as e:
            logger.error(f"Failed to create Excel report: {e}")
            raise

    def get_usage_statistics(self, start_date: date, end_date: date) -> Dict:
        """
        Get usage statistics for a date range.

        Args:
            start_date: Start date
            end_date: End date

        Returns:
            Dictionary with usage statistics
        """
        try:
            # Total items used
            query_total = """
            SELECT SUM(ri.quantity_borrowed) as total_used
            FROM Requisition_Items ri
            JOIN Requisitions r ON r.id = ri.requisition_id
            WHERE r.lab_activity_date BETWEEN ? AND ?
            """
            total_rows = db.execute_query(query_total, (start_date.isoformat(), end_date.isoformat()))
            total_used = total_rows[0]['total_used'] if total_rows and total_rows[0]['total_used'] else 0

            # Items by category
            query_categories = """
            SELECT c.name as category, SUM(ri.quantity_borrowed) as qty
            FROM Requisition_Items ri
            JOIN Requisitions r ON r.id = ri.requisition_id
            JOIN Items i ON i.id = ri.item_id
            JOIN Categories c ON c.id = i.category_id
            WHERE r.lab_activity_date BETWEEN ? AND ?
            GROUP BY c.id, c.name
            ORDER BY qty DESC
            """
            category_rows = db.execute_query(query_categories, (start_date.isoformat(), end_date.isoformat()))

            # Top used items
            query_top_items = """
            SELECT i.name as item_name, SUM(ri.quantity_borrowed) as qty
            FROM Requisition_Items ri
            JOIN Requisitions r ON r.id = ri.requisition_id
            JOIN Items i ON i.id = ri.item_id
            WHERE r.lab_activity_date BETWEEN ? AND ?
            GROUP BY i.id, i.name
            ORDER BY qty DESC
            LIMIT 10
            """
            top_items_rows = db.execute_query(query_top_items, (start_date.isoformat(), end_date.isoformat()))

            return {
                'total_items_used': total_used,
                'categories': category_rows,
                'top_items': top_items_rows,
                'date_range': f"{start_date} to {end_date}"
            }

        except Exception as e:
            logger.error(f"Failed to get usage statistics: {e}")
            return {
                'total_items_used': 0,
                'categories': [],
                'top_items': [],
                'date_range': f"{start_date} to {end_date}"
            }


# Global report generator instance
report_generator = ReportGenerator()

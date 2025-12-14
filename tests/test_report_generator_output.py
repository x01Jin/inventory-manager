from datetime import date, datetime
from pathlib import Path

from inventory_app.database.connection import db
from inventory_app.database.models import Item, Requester, Requisition, RequisitionItem
from inventory_app.gui.reports.report_generator import ReportGenerator
from inventory_app.services.stock_movement_service import StockMovementService

from openpyxl import load_workbook
from inventory_app.gui.reports.report_utils import date_formatter


def setup_temp_db(tmp_path):
    tmp_db_path = tmp_path / "test_inventory.db"
    if tmp_db_path.exists():
        tmp_db_path.unlink()
    db.db_path = tmp_db_path
    assert db.create_database() is True
    return tmp_db_path


def test_usage_report_generates_valid_excel(tmp_path):
    setup_temp_db(tmp_path)

    # Create two items
    pencil = Item()
    pencil.name = "Pencil"
    pencil.category_id = 1
    assert pencil.save(editor_name="tester", batch_quantity=10) is True
    assert pencil.id is not None

    eraser = Item()
    eraser.name = "Eraser"
    eraser.category_id = 1
    assert eraser.save(editor_name="tester", batch_quantity=5) is True
    assert eraser.id is not None

    # Create a requester
    reqr = Requester(name="Student", affiliation="Grade 10", group_name="A")
    assert reqr.save() is True
    assert reqr.id is not None

    # Create two requisitions with items on two consecutive days
    r1 = Requisition()
    assert reqr.id is not None
    r1.requester_id = reqr.id
    r1.expected_request = datetime(2023, 1, 1, 10, 0, 0)
    r1.expected_return = datetime(2023, 1, 2, 10, 0, 0)
    r1.lab_activity_name = "Chem Lab"
    assert r1.save("tester") is True
    assert r1.id is not None

    ri1 = RequisitionItem()
    assert r1.id is not None
    ri1.requisition_id = r1.id
    assert pencil.id is not None
    ri1.item_id = pencil.id
    ri1.quantity_requested = 2
    assert ri1.save() is True

    r2 = Requisition()
    assert reqr.id is not None
    r2.requester_id = reqr.id
    r2.expected_request = datetime(2023, 1, 2, 11, 0, 0)
    r2.expected_return = datetime(2023, 1, 3, 11, 0, 0)
    r2.lab_activity_name = "Bio Lab"
    assert r2.save("tester") is True
    assert r2.id is not None

    ri2 = RequisitionItem()
    assert r2.id is not None
    ri2.requisition_id = r2.id
    assert pencil.id is not None
    ri2.item_id = pencil.id
    ri2.quantity_requested = 3
    assert ri2.save() is True

    ri3 = RequisitionItem()
    assert r2.id is not None
    ri3.requisition_id = r2.id
    assert eraser.id is not None
    ri3.item_id = eraser.id
    ri3.quantity_requested = 1
    assert ri3.save() is True

    # Generate report for the two-day range
    gen = ReportGenerator()
    out = tmp_path / "usage_report_test.xlsx"
    result = gen.generate_report(
        date(2023, 1, 1), date(2023, 1, 2), output_path=str(out)
    )

    assert isinstance(result, str)
    assert result != "" and Path(result).exists()

    wb = load_workbook(result)
    ws = wb.active
    assert ws is not None

    # Headers are on row 4
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    # Should include period headers for the two days and TOTAL QUANTITY
    expected_h1 = date_formatter.format_period_header(date(2023, 1, 1), "daily")
    expected_h2 = date_formatter.format_period_header(date(2023, 1, 2), "daily")
    assert any(expected_h1 in str(h) for h in headers)
    assert any(expected_h2 in str(h) for h in headers)
    assert "Total Quantity" in headers

    # Find Pencil row and assert totals (2 + 3 = 5)
    pencil_row = None
    for r in range(5, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Pencil":
            pencil_row = [
                ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)
            ]
            break
    assert pencil_row is not None
    assert pencil_row[-1] == 5

    # Verify totals row exists and grand total equals 6 (Pencil 5 + Eraser 1)
    total_row_index = ws.max_row
    assert ws.cell(row=total_row_index, column=1).value == "Total"
    assert ws.cell(row=total_row_index, column=ws.max_column).value == 6


def test_low_stock_inventory_report(tmp_path):
    setup_temp_db(tmp_path)

    # Create two items: one low (5), one sufficient (20)
    # create items with larger original stock and then record consumption
    low = Item()
    low.name = "LowItem"
    low.category_id = 1
    assert low.save(editor_name="tester", batch_quantity=100) is True
    assert low.id is not None

    high = Item()
    high.name = "HighItem"
    high.category_id = 1
    assert high.save(editor_name="tester", batch_quantity=20) is True
    assert high.id is not None

    # Consume enough from LowItem to drop it below 20% (100 -> 15 left)
    svc = StockMovementService()
    svc.record_consumption(low.id, 85, None, "consumed for test")

    gen = ReportGenerator()
    out = tmp_path / "inventory_lowstock.xlsx"
    result = gen.generate_inventory_report(
        "Low Stock Alert", date.today(), date.today(), output_path=str(out)
    )

    assert isinstance(result, str)
    assert result != "" and Path(result).exists()

    wb = load_workbook(result)
    ws = wb.active
    assert ws is not None

    # Ensure LowItem present and HighItem absent
    names = [ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)]
    assert "LowItem" in names
    assert "HighItem" not in names


def test_low_stock_threshold_parameter(tmp_path):
    setup_temp_db(tmp_path)

    # Create two items: one low (5), one sufficient (20)
    low = Item()
    low.name = "LowItem"
    low.category_id = 1
    assert low.save(editor_name="tester", batch_quantity=5) is True

    high = Item()
    high.name = "HighItem"
    high.category_id = 1
    assert high.save(editor_name="tester", batch_quantity=20) is True

    gen = ReportGenerator()
    out = tmp_path / "inventory_lowstock_thresh.xlsx"
    # Use a threshold lower than low item to ensure exclusion (threshold=3)
    result = gen.generate_inventory_report(
        "Low Stock Alert",
        date.today(),
        date.today(),
        output_path=str(out),
        low_stock_threshold=3,
    )

    # No rows expected; API returns error message string when no data matches
    assert isinstance(result, str)
    assert "Failed to generate Low Stock Alert" in result

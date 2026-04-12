from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from PyQt6.QtWidgets import QMessageBox, QApplication
from inventory_app.gui.reports.reports_page import ReportsPage
from inventory_app.gui.reports.excel_utils import create_excel_report


def test_reports_page_filters(qtbot):
    """Verify that report filters behave correctly."""
    page = ReportsPage()
    qtbot.addWidget(page)

    # Switch to Usage by Grade Level report first to make filters visible
    idx_report = page.usage_report_type.findData("grade_level")
    page.usage_report_type.setCurrentIndex(idx_report)

    # Check default filter state within that report type
    assert page.usage_filter_type_combo.itemText(0) == "All Grades & Sections"
    assert page.usage_filter_value_combo.isHidden()

    # Switch to Grade Level filter
    idx = page.usage_filter_type_combo.findText("Grade Level")
    page.usage_filter_type_combo.setCurrentIndex(idx)
    assert not page.usage_filter_value_combo.isHidden()
    assert page.usage_filter_value_combo.itemText(0) == "All Grades"


def test_excel_report_ux_features(tmp_path):
    """Verify that generated Excel reports have correct UX features (freeze, filters)."""
    data = [{"Item": "Test Item", "Total Quantity": 10}]
    out = tmp_path / "report.xlsx"
    create_excel_report(data, out, "Test UI Report", date(2025, 1, 1), date(2025, 1, 2))

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None

    # Verify freeze panes (A5 is standard for our reports Header at 4)
    assert ws.freeze_panes == "A5"
    assert ws.auto_filter.ref.startswith("A4")

    # Total row check
    total_row = 5 + len(data)
    assert ws.cell(row=total_row, column=1).value == "Total"


def test_excel_report_column_width_accounts_for_data_content(tmp_path):
    """Column widths should account for long data values, not header text only."""
    long_name = "Sodium Bicarbonate Analytical Grade Very Long Label"
    data = [{"Item": long_name, "Total Quantity": 12}]
    out = tmp_path / "report_width.xlsx"

    create_excel_report(data, out, "Width Test", date(2025, 1, 1), date(2025, 1, 2))

    wb = load_workbook(out)
    ws = wb.active
    assert ws is not None

    # Column A should expand beyond baseline minimum because of long item text.
    assert ws.column_dimensions["A"].width > 20


def test_reports_centralized_labels():
    """Verify that reports page uses centralized ReportConfig labels."""
    from pathlib import Path

    page_source = Path("inventory_app/gui/reports/reports_page.py").read_text(
        encoding="utf-8"
    )

    # Ensure tooltips and labels reference ReportConfig constants
    assert "ReportConfig.GRANULARITY_TOOLTIP" in page_source
    assert "ReportConfig.LABELS" in page_source


def test_trends_category_filter_passed_to_worker(qtbot, monkeypatch):
    """Trends category selector should propagate normalized value into ReportWorker."""

    created_workers = []

    class _DummySignal:
        def connect(self, _slot):
            return None

    class _DummyWorker:
        def __init__(self, _report_type, _start_date, _end_date, **kwargs):
            self.kwargs = kwargs
            self.progress = _DummySignal()
            self.finished = _DummySignal()
            self.error = _DummySignal()
            created_workers.append(self)

        def start(self):
            return None

    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.ReportWorker", _DummyWorker
    )

    page = ReportsPage()
    qtbot.addWidget(page)

    assert page.trends_category_combo.itemText(0) == "All Categories"

    page.trends_category_combo.setCurrentText("All Categories")
    page.generate_trends_report()
    assert created_workers[-1].kwargs.get("category_filter") == ""

    if page.trends_category_combo.findText("Equipment") == -1:
        page.trends_category_combo.addItem("Equipment")
    page.trends_category_combo.setCurrentText("Equipment")
    page.generate_trends_report()
    assert created_workers[-1].kwargs.get("category_filter") == "Equipment"


def test_reports_list_tracks_reports_folder_contents(qtbot, monkeypatch, tmp_path):
    """Generated reports list should mirror current .xlsx files in reports folder."""
    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    first_file = reports_dir / "first.xlsx"
    first_file.write_text("placeholder", encoding="utf-8")

    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.get_reports_directory",
        lambda create=True: reports_dir,
    )
    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.list_report_files",
        lambda: sorted(reports_dir.glob("*.xlsx"), key=lambda p: p.name, reverse=True),
    )

    page = ReportsPage()
    qtbot.addWidget(page)
    qtbot.waitUntil(lambda: len(page._generated_report_paths) == 1, timeout=3000)
    assert Path(page._generated_report_paths[0]).name == "first.xlsx"

    second_file = reports_dir / "second.xlsx"
    second_file.write_text("placeholder", encoding="utf-8")
    page.schedule_report_files_refresh(force=True)

    qtbot.waitUntil(lambda: len(page._generated_report_paths) == 2, timeout=3000)
    assert {Path(path).name for path in page._generated_report_paths} == {
        "first.xlsx",
        "second.xlsx",
    }


def test_copy_selected_report_sets_file_clipboard(qtbot, monkeypatch, tmp_path):
    """Copy action should place selected report as a file URL in clipboard mime data."""
    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    report_file = reports_dir / "copy_me.xlsx"
    report_file.write_text("placeholder", encoding="utf-8")

    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.get_reports_directory",
        lambda create=True: reports_dir,
    )
    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.list_report_files",
        lambda: [report_file],
    )

    page = ReportsPage()
    qtbot.addWidget(page)
    qtbot.waitUntil(lambda: len(page._generated_report_paths) == 1, timeout=3000)
    page.results_list.setCurrentRow(0)

    page.copy_selected_report()
    mime_data = QApplication.clipboard().mimeData()
    assert mime_data is not None
    assert mime_data.hasUrls()
    assert Path(mime_data.urls()[0].toLocalFile()) == report_file


def test_delete_selected_report_requires_confirmation_and_audit(
    qtbot, monkeypatch, tmp_path
):
    """Delete action should remove file and log REPORT_DELETED with provided editor."""
    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    report_file = reports_dir / "delete_me.xlsx"
    report_file.write_text("placeholder", encoding="utf-8")

    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.get_reports_directory",
        lambda create=True: reports_dir,
    )
    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.list_report_files",
        lambda: [report_file] if report_file.exists() else [],
    )
    monkeypatch.setattr(
        QMessageBox,
        "question",
        lambda *_args, **_kwargs: QMessageBox.StandardButton.Yes,
    )
    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.QInputDialog.getText",
        lambda *_args, **_kwargs: ("QA", True),
    )

    activity_calls = []

    def _capture_activity(*args, **kwargs):
        activity_calls.append((args, kwargs))
        return True

    monkeypatch.setattr(
        "inventory_app.gui.reports.reports_page.activity_logger.log_activity",
        _capture_activity,
    )

    page = ReportsPage()
    qtbot.addWidget(page)
    qtbot.waitUntil(lambda: len(page._generated_report_paths) == 1, timeout=3000)
    page.results_list.setCurrentRow(0)

    page.delete_selected_report()

    assert not report_file.exists()
    assert activity_calls
    assert activity_calls[0][0][0] == "REPORT_DELETED"
    assert activity_calls[0][1].get("user_name") == "QA"

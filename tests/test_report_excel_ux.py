from datetime import date

from openpyxl import load_workbook

from inventory_app.gui.reports.excel_utils import create_excel_report


def test_create_excel_report_freeze_filter_and_totals(tmp_path):
    # Minimal data set
    data = [
        {"Item": "Widget A", "Total Quantity": 10},
        {"Item": "Widget B", "Total Quantity": 5},
    ]
    output_file = tmp_path / "test_report.xlsx"
    start_date = date(2025, 1, 1)
    end_date = date(2025, 1, 31)

    create_excel_report(data, output_file, "Test Report", start_date, end_date)

    # Open workbook and check attributes
    wb = load_workbook(output_file)
    ws = wb.active
    assert ws is not None

    # Freeze panes should be set to A5
    assert ws.freeze_panes == "A5"

    # Auto filter should start at A4 and include header row and data rows
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref.startswith("A4")

    # Check numeric formatting for quantity column (column 2)
    formatted_cell = ws.cell(row=5, column=2)
    assert formatted_cell.number_format == "#,##0"

    # Check totals row
    totals_row = 5 + len(data)
    assert ws.cell(row=totals_row, column=1).value == "Total"


def test_create_excel_report_respects_manual_granularity_in_headers(tmp_path):
    # Data with a monthly period key; date range is short but we force monthly
    data = [{"Item": "Widget A", "2023-01": 10}]
    output_file = tmp_path / "test_report_monthly.xlsx"
    start_date = date(2023, 1, 1)
    end_date = date(2023, 1, 7)

    create_excel_report(
        data, output_file, "Test Report", start_date, end_date, granularity="monthly"
    )

    wb = load_workbook(output_file)
    ws = wb.active
    assert ws is not None

    # Header row (row 4), find monthly header text
    headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
    assert any("Jan" in str(h) for h in headers)

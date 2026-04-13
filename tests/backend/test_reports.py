import pytest
from datetime import date
from pathlib import Path
from openpyxl import load_workbook
from inventory_app.database.connection import db
from inventory_app.database.models import Item
from inventory_app.gui.reports.report_generator import ReportGenerator, report_generator
from inventory_app.utils.activity_logger import activity_logger
from inventory_app.gui.reports.query_builder import ReportQueryBuilder
from inventory_app.gui.reports.data_sources import (
    get_defective_items_data,
    get_stock_levels_data,
    get_audit_log_data,
    get_update_history_data,
    get_usage_by_grade_level_data,
    get_trends_data,
    get_dynamic_report_data,
)
from inventory_app.gui.reports.monthly_usage_report import generate_monthly_usage_report
from inventory_app.gui.reports.report_utils import date_formatter
from inventory_app.gui.reports.report_worker import ReportWorker


@pytest.fixture
def temp_db(tmp_path):
    """Create a temporary database for each test and ensure schema is applied."""
    db_file = tmp_path / "test_inventory.db"
    db.db_path = db_file
    assert db.create_database() is True
    yield db_file


def test_query_builder_logic():
    """Verify that ReportQueryBuilder correctly generates SQL and parameters."""
    builder = ReportQueryBuilder()
    start = date(2025, 1, 1)
    end = date(2025, 1, 5)

    # Test periodic query building
    query, params = builder.build_dynamic_report_query(start, end, "daily")

    # Verify parameterization (no string injection)
    assert query.count("?") == len(params)
    assert "r.lab_activity_date" in query

    # Verify period key generation
    from inventory_app.gui.reports.report_utils import date_formatter

    keys = date_formatter.get_period_keys(start, end, "daily")
    assert len(keys) == 5
    for k in keys:
        assert f'"{k}"' in query


def test_usage_report_generation(temp_db, tmp_path):
    """Verify end-to-call usage report generation to Excel."""
    # Setup data
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("Pencil", 1),
        return_last_id=True,
    )[1]
    reqr_id = db.execute_update(
        "INSERT INTO Requesters (name) VALUES (?)", ("Student",), return_last_id=True
    )[1]

    # Requisition on 2025-01-01
    req_id = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            reqr_id,
            "requested",
            "2025-01-01",
            "Test Activity",
            "2025-01-01 09:00:00",
            "2025-01-01 12:00:00",
        ),
        return_last_id=True,
    )[1]
    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (req_id, item_id, 5),
    )

    gen = ReportGenerator()
    out_file = tmp_path / "usage.xlsx"
    result = gen.generate_report(
        date(2025, 1, 1), date(2025, 1, 1), output_path=str(out_file)
    )

    assert isinstance(result, str)
    assert Path(result).exists()
    wb = load_workbook(result)
    ws = wb.active
    assert ws is not None

    # Find Pencil row and verify total
    found = False
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "Pencil":
            # Total is usually the last column
            assert ws.cell(row=row, column=ws.max_column).value == 5
            found = True
            break
    assert found


def test_usage_report_default_output_directory(temp_db, tmp_path, monkeypatch):
    """Default report output path should be created inside the dedicated reports folder."""
    monkeypatch.chdir(tmp_path)

    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("Marker", 1),
        return_last_id=True,
    )[1]
    reqr_id = db.execute_update(
        "INSERT INTO Requesters (name) VALUES (?)", ("User",), return_last_id=True
    )[1]
    req_id = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            reqr_id,
            "requested",
            "2025-01-01",
            "Activity",
            "2025-01-01 09:00:00",
            "2025-01-01 12:00:00",
        ),
        return_last_id=True,
    )[1]
    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (req_id, item_id, 2),
    )

    gen = ReportGenerator()
    result = gen.generate_report(date(2025, 1, 1), date(2025, 1, 1))

    assert isinstance(result, str)
    output_path = Path(result)
    assert output_path.exists()
    assert output_path.parent == tmp_path / "reports"


def test_report_date_range_description_uses_explicit_dates():
    """Report period header should show exact selected date range."""
    desc = date_formatter.get_date_range_description(date(2026, 1, 1), date(2026, 1, 5))
    assert desc == "Jan 1, 2026 - Jan 5, 2026"

    single = date_formatter.get_date_range_description(
        date(2026, 1, 1), date(2026, 1, 1)
    )
    assert single == "Jan 1, 2026"


def test_report_worker_path_classification_rejects_failure_payloads():
    """Worker should only treat xlsx file paths as successful completion values."""
    assert ReportWorker._is_successful_path("report_20260101.xlsx") is True
    assert (
        ReportWorker._is_successful_path(" Failed to generate Audit Log Report")
        is False
    )
    assert ReportWorker._is_successful_path("Unknown report type: audit") is False
    assert ReportWorker._is_successful_path("") is False


def test_specialized_reports_data_retrieval(temp_db):
    """Verify data retrieval for specialized reports (Disposal, Defective)."""
    # Setup defective item
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("Broken Item", 1),
        return_last_id=True,
    )[1]
    # Setup defective item linked to a requisition
    reqr_id = db.execute_update(
        "INSERT INTO Requesters (name) VALUES (?)", ("Bob",), return_last_id=True
    )[1]
    req_id = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            reqr_id,
            "returned",
            "2025-01-01",
            "Returned Activity",
            "2025-01-01 09:00:00",
            "2025-01-01 12:00:00",
        ),
        return_last_id=True,
    )[1]

    db.execute_update(
        "INSERT INTO Defective_Items (item_id, requisition_id, quantity, notes, reported_by, reported_date) VALUES (?, ?, ?, ?, ?, ?)",
        (item_id, req_id, 1, "Cracked", "tester", "2025-01-01"),
    )

    data = get_defective_items_data(date(2025, 1, 1), date(2025, 1, 1))
    assert len(data) == 1
    assert data[0]["Item Name"] == "Broken Item"
    assert data[0]["Notes"] == "Cracked"


def test_update_history_report_integrity(temp_db):
    """Verify that update history is correctly logged and retrieved."""
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("HistoryItem", 1),
        return_last_id=True,
    )[1]

    # Simulate update via model to trigger history logging
    assert item_id is not None
    item = Item.get_by_id(item_id)
    assert item is not None
    item.name = "UpdatedName"
    assert item.save(editor_name="RefactorTester") is True

    # Check history table
    history = db.execute_query(
        "SELECT * FROM Update_History WHERE item_id = ?", (item_id,)
    )
    assert len(history) >= 1
    assert history[0]["editor_name"] == "RefactorTester"
    assert any(row.get("field_name") == "name" for row in history)
    matching = [row for row in history if row.get("field_name") == "name"]
    assert matching[0]["old_value"] == "HistoryItem"
    assert matching[0]["new_value"] == "UpdatedName"

    report_rows = get_update_history_data(date(2020, 1, 1), date(2100, 1, 1))
    target = next(
        (row for row in report_rows if row.get("Item Name") == "UpdatedName"), None
    )
    assert target is not None
    assert target["PO Number"] == "N/A"


def test_audit_log_data_source(temp_db):
    """Unified audit log dataset should include history and activity records."""
    created_item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("Audit Item", 1),
        return_last_id=True,
    )[1]
    assert created_item_id is not None
    item_id = created_item_id

    item = Item.get_by_id(item_id)
    assert item is not None
    item.name = "Audit Item Updated"
    assert item.save(editor_name="AuditTester") is True

    rows = get_audit_log_data(
        start_date=date(2020, 1, 1),
        end_date=date(2100, 1, 1),
        editor_filter="AuditTester",
    )

    assert rows
    assert any(r.get("Action") == "ITEM_UPDATE" for r in rows)
    assert any(r.get("Editor") == "AuditTester" for r in rows)
    assert any(" at " in (r.get("Timestamp") or "") for r in rows)


def test_audit_log_groups_field_level_item_updates(temp_db):
    """Audit report should merge same-event item field updates into one readable row."""
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("Grouped Audit Item", 1),
        return_last_id=True,
    )[1]
    timestamp = "2026-04-13 10:15:00"

    db.execute_update(
        "INSERT INTO Update_History (item_id, editor_name, reason, edit_timestamp, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (item_id, "Editor", "Item field updated", timestamp, "brand", "NA", "LabCorp"),
    )
    db.execute_update(
        "INSERT INTO Update_History (item_id, editor_name, reason, edit_timestamp, field_name, old_value, new_value) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (item_id, "Editor", "Item field updated", timestamp, "supplier_id", "", "8"),
    )

    rows = get_audit_log_data(
        start_date=date(2026, 4, 13),
        end_date=date(2026, 4, 13),
        editor_filter="Editor",
        action_filter="ITEM_UPDATE",
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["Action"] == "ITEM_UPDATE"
    assert row["Editor"] == "Editor"
    assert "Updated Grouped Audit Item" in row["Summary"]
    assert "Brand: NA -> LabCorp" in row["Change Details"]
    assert "Supplier ID: (empty) -> 8" in row["Change Details"]


def test_recent_activity_order_handles_mixed_timestamp_formats(temp_db):
    """Recent activity should sort by actual datetime even with mixed timestamp string formats."""
    activity_logger.log_activity(
        "TEST_ACTION",
        "older event",
        timestamp="2026-04-13 08:00:00",
    )
    activity_logger.log_activity(
        "TEST_ACTION",
        "newer event",
        timestamp="2026-04-13T09:00:00+00:00",
    )

    activities = activity_logger.get_recent_activities(2)
    assert len(activities) == 2
    assert activities[0]["description"] == "newer event"
    assert activities[1]["description"] == "older event"


def test_audit_log_includes_defective_confirmation_activity(temp_db):
    """Unified audit data should include defective confirmation actions from Activity_Log."""
    activity_logger.log_activity(
        activity_logger.DEFECTIVE_NOT_DEFECTIVE,
        "marked quantity as not defective 1 for item_id=9",
        entity_id=9,
        entity_type="item",
        user_name="auditor",
        timestamp="2026-04-13T11:00:00+00:00",
    )

    rows = get_audit_log_data(
        start_date=date(2026, 4, 13),
        end_date=date(2026, 4, 13),
        action_filter="DEFECTIVE_NOT_DEFECTIVE",
    )

    assert rows
    assert rows[0]["Action"] == "DEFECTIVE_NOT_DEFECTIVE"
    assert rows[0]["Editor"] == "auditor"


def test_task10_stock_levels_data_policy(temp_db):
    """Task 10: stock report reflects consumable depletion and non-consumable disposal only."""
    consumable_id = db.execute_update(
        "INSERT INTO Items (name, category_id, is_consumable) VALUES (?, ?, ?)",
        ("Report Consumable", 1, 1),
        return_last_id=True,
    )[1]
    non_consumable_id = db.execute_update(
        "INSERT INTO Items (name, category_id, is_consumable) VALUES (?, ?, ?)",
        ("Report NonConsumable", 1, 0),
        return_last_id=True,
    )[1]

    consumable_batch_id = db.execute_update(
        "INSERT INTO Item_Batches (item_id, batch_number, quantity_received, date_received) VALUES (?, ?, ?, ?)",
        (consumable_id, 1, 50, "2025-01-01"),
        return_last_id=True,
    )[1]
    non_consumable_batch_id = db.execute_update(
        "INSERT INTO Item_Batches (item_id, batch_number, quantity_received, date_received) VALUES (?, ?, ?, ?)",
        (non_consumable_id, 1, 20, "2025-01-01"),
        return_last_id=True,
    )[1]

    # Consumable final stock: 50 - 12 - 3 + 2 = 37
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (consumable_id, consumable_batch_id, "CONSUMPTION", 12, "2025-01-02"),
    )
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (consumable_id, consumable_batch_id, "DISPOSAL", 3, "2025-01-03"),
    )
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (consumable_id, consumable_batch_id, "RETURN", 2, "2025-01-04"),
    )

    # Non-consumable final stock: 20 - 5 = 15 (REQUEST/RETURN must not change stock)
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (non_consumable_id, non_consumable_batch_id, "REQUEST", 7, "2025-01-02"),
    )
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (non_consumable_id, non_consumable_batch_id, "RETURN", 7, "2025-01-03"),
    )
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (non_consumable_id, non_consumable_batch_id, "DISPOSAL", 5, "2025-01-05"),
    )

    rows = get_stock_levels_data()
    by_name = {row["Item Name"]: row for row in rows}

    assert by_name["Report Consumable"]["Current Stock"] == 37
    assert by_name["Report NonConsumable"]["Current Stock"] == 15


def test_usage_by_grade_level_data_task14_aggregation(temp_db):
    """Task 14: Grade-level report includes Grade 7-10 tallies and Task 10 stock semantics."""
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id, is_consumable) VALUES (?, ?, ?)",
        ("Grade Tally Item", 1, 1),
        return_last_id=True,
    )[1]
    batch_id = db.execute_update(
        "INSERT INTO Item_Batches (item_id, batch_number, quantity_received, date_received) VALUES (?, ?, ?, ?)",
        (item_id, 1, 100, "2025-01-01"),
        return_last_id=True,
    )[1]

    # Consumable stock: 100 - 10 - 2 + 1 = 89
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (item_id, batch_id, "CONSUMPTION", 10, "2025-01-05"),
    )
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (item_id, batch_id, "DISPOSAL", 2, "2025-01-06"),
    )
    db.execute_update(
        "INSERT INTO Stock_Movements (item_id, batch_id, movement_type, quantity, movement_date) VALUES (?, ?, ?, ?, ?)",
        (item_id, batch_id, "RETURN", 1, "2025-01-07"),
    )

    reqr_g7 = db.execute_update(
        "INSERT INTO Requesters (name, grade_level, section) VALUES (?, ?, ?)",
        ("Grade7 User", "Grade 7", "A"),
        return_last_id=True,
    )[1]
    reqr_g8 = db.execute_update(
        "INSERT INTO Requesters (name, grade_level, section) VALUES (?, ?, ?)",
        ("Grade8 User", "Grade 8", "B"),
        return_last_id=True,
    )[1]

    req_non_individual = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return, is_individual) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (
            reqr_g7,
            "requested",
            "2025-01-08",
            "Class Activity",
            "2025-01-08 09:00:00",
            "2025-01-08 12:00:00",
            0,
        ),
        return_last_id=True,
    )[1]
    req_individual = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return, is_individual) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (
            reqr_g8,
            "requested",
            "2025-01-09",
            "Individual Activity",
            "2025-01-09 09:00:00",
            "2025-01-09 12:00:00",
            1,
        ),
        return_last_id=True,
    )[1]

    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (req_non_individual, item_id, 4),
    )
    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (req_individual, item_id, 3),
    )

    rows_all = get_usage_by_grade_level_data(
        date(2025, 1, 1),
        date(2025, 1, 31),
        show_individual_only=False,
    )
    assert rows_all
    row = rows_all[0]
    assert row["ACTUAL INVENTORY"] == 89
    assert row["GRADE 7"] == 4
    assert row["GRADE 8"] == 3
    assert row["GRADE 9"] == 0
    assert row["GRADE 10"] == 0
    assert row["TOTAL QUANTITY"] == 7

    rows_individual = get_usage_by_grade_level_data(
        date(2025, 1, 1),
        date(2025, 1, 31),
        show_individual_only=True,
    )
    assert rows_individual
    individual_row = rows_individual[0]
    assert individual_row["GRADE 7"] == 0
    assert individual_row["GRADE 8"] == 3
    assert individual_row["TOTAL QUANTITY"] == 3


def test_monthly_usage_report_includes_grade_tally_columns(temp_db, tmp_path):
    """Task 14: Monthly export includes Grade 7-10 tally columns with exact headers."""
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id, is_consumable) VALUES (?, ?, ?)",
        ("Monthly Grade Item", 1, 1),
        return_last_id=True,
    )[1]
    db.execute_update(
        "INSERT INTO Item_Batches (item_id, batch_number, quantity_received, date_received) VALUES (?, ?, ?, ?)",
        (item_id, 1, 20, "2025-01-01"),
    )
    reqr_id = db.execute_update(
        "INSERT INTO Requesters (name, grade_level, section) VALUES (?, ?, ?)",
        ("Monthly Grade7", "Grade 7", "A"),
        return_last_id=True,
    )[1]
    req_id = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            reqr_id,
            "requested",
            "2025-01-10",
            "Monthly Activity",
            "2025-01-10 09:00:00",
            "2025-01-10 12:00:00",
        ),
        return_last_id=True,
    )[1]
    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (req_id, item_id, 2),
    )

    out_file = tmp_path / "monthly_usage.xlsx"
    result = generate_monthly_usage_report(2025, 1, output_path=str(out_file))
    assert isinstance(result, str)
    assert Path(result).exists()

    wb = load_workbook(result)
    ws = wb.active
    assert ws is not None

    assert ws.cell(row=4, column=6).value == "PO NUMBER"
    assert ws.cell(row=4, column=8).value == "GRADE 7"
    assert ws.cell(row=4, column=9).value == "GRADE 8"
    assert ws.cell(row=4, column=10).value == "GRADE 9"
    assert ws.cell(row=4, column=11).value == "GRADE 10"
    assert ws.cell(row=4, column=12).value == "TOTAL GRADE USAGE"


def test_trends_data_respects_category_filter(temp_db):
    """Trends data should include only rows from selected category."""
    categories = db.execute_query(
        "SELECT id, name FROM Categories WHERE name IN (?, ?)",
        ("Equipment", "Apparatus"),
    )
    category_map = {row["name"]: row["id"] for row in categories}
    assert "Equipment" in category_map
    assert "Apparatus" in category_map

    equipment_item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("Equipment Trend Item", category_map["Equipment"]),
        return_last_id=True,
    )[1]
    apparatus_item_id = db.execute_update(
        "INSERT INTO Items (name, category_id) VALUES (?, ?)",
        ("Apparatus Trend Item", category_map["Apparatus"]),
        return_last_id=True,
    )[1]

    requester_id = db.execute_update(
        "INSERT INTO Requesters (name) VALUES (?)",
        ("Trend Requester",),
        return_last_id=True,
    )[1]

    req_one = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            requester_id,
            "requested",
            "2025-01-10",
            "Equipment Activity",
            "2025-01-10 09:00:00",
            "2025-01-10 12:00:00",
        ),
        return_last_id=True,
    )[1]
    req_two = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            requester_id,
            "requested",
            "2025-01-12",
            "Apparatus Activity",
            "2025-01-12 09:00:00",
            "2025-01-12 12:00:00",
        ),
        return_last_id=True,
    )[1]

    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (req_one, equipment_item_id, 4),
    )
    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (req_two, apparatus_item_id, 3),
    )

    rows = get_trends_data(
        date(2025, 1, 1),
        date(2025, 1, 31),
        granularity="daily",
        group_by="item",
        category_filter="Equipment",
    )

    assert rows
    assert all(row.get("CATEGORIES") == "Equipment" for row in rows)
    assert any(row.get("ITEMS") == "Equipment Trend Item" for row in rows)
    assert all(row.get("ITEMS") != "Apparatus Trend Item" for row in rows)


def test_usage_report_supplier_filter_targets_item_supplier(temp_db):
    """Supplier filter should target item supplier names, not requester fields."""
    supplier_a = db.execute_update(
        "INSERT INTO Suppliers (name) VALUES (?)",
        ("Supplier A",),
        return_last_id=True,
    )[1]
    supplier_b = db.execute_update(
        "INSERT INTO Suppliers (name) VALUES (?)",
        ("Supplier B",),
        return_last_id=True,
    )[1]

    item_a = db.execute_update(
        "INSERT INTO Items (name, category_id, supplier_id) VALUES (?, ?, ?)",
        ("Supplier A Item", 1, supplier_a),
        return_last_id=True,
    )[1]
    item_b = db.execute_update(
        "INSERT INTO Items (name, category_id, supplier_id) VALUES (?, ?, ?)",
        ("Supplier B Item", 1, supplier_b),
        return_last_id=True,
    )[1]

    requester_id = db.execute_update(
        "INSERT INTO Requesters (name, department) VALUES (?, ?)",
        ("Requester", "Not a Supplier"),
        return_last_id=True,
    )[1]

    requisition_id = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            requester_id,
            "requested",
            "2025-02-01",
            "Supplier Filter Activity",
            "2025-02-01 09:00:00",
            "2025-02-01 12:00:00",
        ),
        return_last_id=True,
    )[1]

    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (requisition_id, item_a, 5),
    )
    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (requisition_id, item_b, 3),
    )

    rows = get_dynamic_report_data(
        start_date=date(2025, 2, 1),
        end_date=date(2025, 2, 1),
        granularity="daily",
        supplier_filter="Supplier A",
    )

    assert rows
    assert len(rows) == 1
    assert rows[0]["ITEMS"] == "Supplier A Item"
    assert rows[0]["SUPPLIER"] == "Supplier A"


def test_dynamic_report_data_includes_po_number(temp_db):
    """Dynamic usage report rows should include PO NUMBER when present on the item."""
    requester_id = db.execute_update(
        "INSERT INTO Requesters (name) VALUES (?)",
        ("PO Requester",),
        return_last_id=True,
    )[1]
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id, po_number) VALUES (?, ?, ?)",
        ("PO Usage Item", 1, "PO-DYN-900"),
        return_last_id=True,
    )[1]
    requisition_id = db.execute_update(
        "INSERT INTO Requisitions (requester_id, status, lab_activity_date, lab_activity_name, expected_request, expected_return) VALUES (?, ?, ?, ?, ?, ?)",
        (
            requester_id,
            "requested",
            "2025-02-02",
            "PO Usage Activity",
            "2025-02-02 09:00:00",
            "2025-02-02 12:00:00",
        ),
        return_last_id=True,
    )[1]
    db.execute_update(
        "INSERT INTO Requisition_Items (requisition_id, item_id, quantity_requested) VALUES (?, ?, ?)",
        (requisition_id, item_id, 2),
    )

    rows = get_dynamic_report_data(
        start_date=date(2025, 2, 2),
        end_date=date(2025, 2, 2),
        granularity="daily",
    )
    row = next((entry for entry in rows if entry.get("ITEMS") == "PO Usage Item"), None)

    assert row is not None
    assert row["PO NUMBER"] == "PO-DYN-900"


def test_stock_levels_data_includes_supplier_column(temp_db):
    """Stock levels data should include supplier name for Task 14 output completeness."""
    supplier_id = db.execute_update(
        "INSERT INTO Suppliers (name) VALUES (?)",
        ("Stock Supplier",),
        return_last_id=True,
    )[1]
    item_id = db.execute_update(
        "INSERT INTO Items (name, category_id, supplier_id, po_number, is_consumable) VALUES (?, ?, ?, ?, ?)",
        ("Stock Supplier Item", 1, supplier_id, "PO-STOCK-123", 1),
        return_last_id=True,
    )[1]
    batch_id = db.execute_update(
        "INSERT INTO Item_Batches (item_id, batch_number, quantity_received, date_received) VALUES (?, ?, ?, ?)",
        (item_id, 1, 25, "2025-02-01"),
        return_last_id=True,
    )[1]

    assert batch_id is not None

    rows = get_stock_levels_data()
    row = next(
        (entry for entry in rows if entry.get("Item Name") == "Stock Supplier Item"),
        None,
    )

    assert row is not None
    assert row["Supplier"] == "Stock Supplier"
    assert row["PO Number"] == "PO-STOCK-123"


def test_report_worker_passes_trends_category_filter(monkeypatch):
    """ReportWorker trends path should pass category_filter through to generator."""
    captured = {}

    def _fake_generate_trends_report(*_args, **kwargs):
        captured.update(kwargs)
        return "ok.xlsx"

    monkeypatch.setattr(
        report_generator,
        "generate_trends_report",
        _fake_generate_trends_report,
    )

    worker = ReportWorker(
        "trends",
        date(2025, 1, 1),
        date(2025, 1, 31),
        granularity="weekly",
        category_filter="Equipment",
        include_consumables=True,
    )

    result = worker._generate_trends_report()

    assert result == "ok.xlsx"
    assert captured.get("category_filter") == "Equipment"

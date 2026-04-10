import pytest
from openpyxl import Workbook
from inventory_app.database.connection import db
from inventory_app.services.item_importer import (
    collect_consumable_rows_missing_unit,
    import_items_from_excel,
)
from inventory_app.utils.stock_parser import parse_stock_value, parse_stock_quantity


@pytest.fixture
def temp_db(tmp_path):
    """Create a temporary database for each test and ensure schema is applied."""
    db_file = tmp_path / "test_inventory.db"
    db.db_path = db_file
    assert db.create_database() is True
    yield db_file


def test_stock_string_parsing():
    """Verify parsing logic for stock quantities and sizes from strings."""
    # Simple numeric
    assert parse_stock_quantity(10) == 10
    assert parse_stock_quantity("5") == 5

    # Combined quantity and size
    info = parse_stock_value("900ml")
    assert info["quantity"] == 900
    assert info["size"] == "900ml"

    info = parse_stock_value("2.5 L")
    assert info["quantity"] == 2500
    assert info["size"] == "2.5 L"

    info = parse_stock_value("1 kilo")
    assert info["quantity"] == 1000
    assert info["size"] == "1 kilo"

    info = parse_stock_value("125 gms")
    assert info["quantity"] == 125
    assert info["size"] == "125 gms"

    info = parse_stock_value("1.1 gal")
    assert info["quantity"] == 1100
    assert info["size"] == "1.1 gal"

    info = parse_stock_value("1.1 galon")
    assert info["quantity"] == 1100
    assert info["size"] == "1.1 galon"

    info = parse_stock_value("10 boxes")
    assert info["quantity"] == 10
    assert info["size"] is None

    info = parse_stock_value("1 box (100pcs)")
    assert info["quantity"] == 100
    assert info["size"] is None
    assert info["notes"] == "(100pcs)"

    info = parse_stock_value("2 packs of 50 pcs")
    assert info["quantity"] == 100
    assert info["size"] is None

    # Invalid strings
    with pytest.raises(ValueError):
        parse_stock_value("just text")


def test_excel_importer_logic(temp_db, tmp_path):
    """Verify that importing from Excel correctly populates the database."""
    # Create temp excel file
    excel_path = tmp_path / "test_import.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "stocks", "item type", "category", "supplier"])
    ws.append(["Item A", "1", "Consumables", "Chemicals-Solid", "Malcor Chemicals"])
    ws.append(
        ["Item B", "500ml", "TA, Consumables", "Chemicals-Liquid", "ATR Trading System"]
    )
    ws.append(
        ["Item C", "5", "TA, non consumable", "Equipment", "RNJ Medical Equipment"]
    )
    ws.append(["Item D", "1 box (100pcs)", "Consumables", "Consumables", "Dalkem"])
    ws.append(["Item E", "1 kilo", "Consumables", "Chemicals-Solid", "Instruchem Inc."])
    ws.append(
        [
            "Item F",
            "2.5 L",
            "Consumables",
            "Chemicals-Liquid",
            "Brightway Trading School",
        ]
    )
    wb.save(excel_path)

    # Run import
    imported_count, messages = import_items_from_excel(
        str(excel_path), editor_name="tester"
    )

    assert imported_count == 6

    # Verify DB state
    rows = db.execute_query(
        """
        SELECT i.name, i.is_consumable, i.item_type, i.size, s.name AS supplier_name
        FROM Items i
        LEFT JOIN Suppliers s ON s.id = i.supplier_id
        ORDER BY i.name
        """
    )
    assert len(rows) == 6

    # Item A: Consumable = 1
    assert rows[0]["name"] == "Item A"
    assert rows[0]["is_consumable"] == 1
    assert rows[0]["item_type"] == "Consumable"
    assert rows[0]["supplier_name"] == "Malcor Chemicals"

    # Item B: Size extracted from stocks string
    assert rows[1]["name"] == "Item B"
    assert rows[1]["size"] == "500 mL"
    assert rows[1]["item_type"] == "Consumable"
    assert rows[1]["supplier_name"] == "ATR Trading System"

    # Item B: Batch quantity should use the numeric part from stocks (500ml -> 500)
    item_b_qty = db.execute_query(
        """
        SELECT ib.quantity_received
        FROM Item_Batches ib
        JOIN Items i ON i.id = ib.item_id
        WHERE i.name = ?
        ORDER BY ib.id DESC
        LIMIT 1
        """,
        ("Item B",),
    )
    assert item_b_qty
    assert item_b_qty[0]["quantity_received"] == 500

    # Item C: TA non-consumable (Not consumable)
    assert rows[2]["name"] == "Item C"
    assert rows[2]["is_consumable"] == 0
    assert rows[2]["item_type"] == "TA, non-consumable"
    assert rows[2]["supplier_name"] == "RNJ Medical Equipment"

    # Item D: Boxed piece count converts to usable quantity.
    item_d_qty = db.execute_query(
        """
        SELECT ib.quantity_received
        FROM Item_Batches ib
        JOIN Items i ON i.id = ib.item_id
        WHERE i.name = ?
        ORDER BY ib.id DESC
        LIMIT 1
        """,
        ("Item D",),
    )
    assert item_d_qty
    assert item_d_qty[0]["quantity_received"] == 100

    item_d_specs = db.execute_query(
        "SELECT other_specifications FROM Items WHERE name = ?",
        ("Item D",),
    )
    assert item_d_specs
    assert "(100pcs)" in (item_d_specs[0]["other_specifications"] or "")

    # Item E: Alternate metric alias in stocks should populate size and stock quantity.
    item_e = db.execute_query("SELECT size FROM Items WHERE name = ?", ("Item E",))
    assert item_e
    assert item_e[0]["size"] == "1 kg"

    item_e_qty = db.execute_query(
        """
        SELECT ib.quantity_received
        FROM Item_Batches ib
        JOIN Items i ON i.id = ib.item_id
        WHERE i.name = ?
        ORDER BY ib.id DESC
        LIMIT 1
        """,
        ("Item E",),
    )
    assert item_e_qty
    assert item_e_qty[0]["quantity_received"] == 1000

    # Item F: Liter values should be converted to ml-equivalent usable units.
    item_f = db.execute_query("SELECT size FROM Items WHERE name = ?", ("Item F",))
    assert item_f
    assert item_f[0]["size"] == "2.5 L"

    item_f_qty = db.execute_query(
        """
        SELECT ib.quantity_received
        FROM Item_Batches ib
        JOIN Items i ON i.id = ib.item_id
        WHERE i.name = ?
        ORDER BY ib.id DESC
        LIMIT 1
        """,
        ("Item F",),
    )
    assert item_f_qty
    assert item_f_qty[0]["quantity_received"] == 2500


def test_importer_edge_cases(temp_db, tmp_path):
    """Verify importer behavior with duplicate items and missing categories."""
    excel_path = tmp_path / "edge_cases.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "stocks", "item type", "category"])

    # Duplicate items: Should be imported as separate batches or update existing?
    # Based on current implementation, it likely creates separate items if names are exactly same
    # or fails if there is a unique constraint (which there isn't in schema).
    ws.append(["Duplicate", "10", "Consumables", "Consumables"])
    ws.append(["Duplicate", "20", "Consumables", "Consumables"])

    # Missing category: Should default to 'Uncategorized'
    ws.append(["NoCat", "5", "Consumables", ""])

    wb.save(excel_path)

    imported_count, messages = import_items_from_excel(
        str(excel_path), editor_name="tester"
    )
    assert imported_count == 3

    # Verify Uncategorized
    nocat = db.execute_query("SELECT category_id FROM Items WHERE name = 'NoCat'")
    assert len(nocat) == 1

    # Get Uncategorized ID
    uncat_id = db.execute_query(
        "SELECT id FROM Categories WHERE name = 'Uncategorized'"
    )[0]["id"]
    assert nocat[0]["category_id"] == uncat_id


def test_collect_consumable_rows_missing_unit_detects_decimal_only(temp_db, tmp_path):
    """Detect consumable rows that have decimal stocks without explicit unit text."""
    excel_path = tmp_path / "missing_unit_scan.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "stocks", "item type", "category"])
    ws.append(["Ethanol", "1.5", "Consumable", "Chemicals-Liquid"])
    ws.append(["Ethyl Alcohol 95%", "1.1 gal", "Consumable", "Chemicals-Liquid"])
    ws.append(["Gloves", "100", "Consumable", "Consumables"])
    wb.save(excel_path)

    flagged = collect_consumable_rows_missing_unit(str(excel_path))
    assert len(flagged) == 1
    assert flagged[0]["name"] == "Ethanol"
    assert flagged[0]["stocks"] == "1.5"


def test_importer_applies_unit_overrides_and_skip_rows(temp_db, tmp_path):
    """Apply user unit selections to rows and skip selected ambiguous rows."""
    excel_path = tmp_path / "missing_unit_resolution.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "stocks", "item type", "category"])
    ws.append(["Ethanol", "1.5", "Consumable", "Chemicals-Liquid"])
    ws.append(["Acetone", "2.0", "Consumable", "Chemicals-Liquid"])
    ws.append(["Beaker", "5", "Non-Consumable", "Equipment"])
    wb.save(excel_path)

    flagged = collect_consumable_rows_missing_unit(str(excel_path))
    by_name = {entry["name"]: entry["row_index"] for entry in flagged}

    imported_count, messages = import_items_from_excel(
        str(excel_path),
        editor_name="tester",
        row_unit_overrides={int(by_name["Ethanol"]): "L"},
        rows_to_skip={int(by_name["Acetone"])},
    )

    assert imported_count == 2
    assert any("skipped by user" in msg.lower() for msg in messages)

    ethanol = db.execute_query(
        "SELECT size FROM Items WHERE name = ?",
        ("Ethanol",),
    )
    assert ethanol
    assert ethanol[0]["size"] == "1.5 L"

    ethanol_qty = db.execute_query(
        """
        SELECT ib.quantity_received
        FROM Item_Batches ib
        JOIN Items i ON i.id = ib.item_id
        WHERE i.name = ?
        ORDER BY ib.id DESC
        LIMIT 1
        """,
        ("Ethanol",),
    )
    assert ethanol_qty
    assert ethanol_qty[0]["quantity_received"] == 1500

    acetone = db.execute_query("SELECT id FROM Items WHERE name = ?", ("Acetone",))
    assert not acetone


def test_importer_normalizes_spaced_category_aliases(temp_db, tmp_path):
    """Importer should normalize spaced chemical category labels to canonical names."""
    excel_path = tmp_path / "category_alias_import.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "stocks", "item type", "category"])
    ws.append(["AliasChem", "50", "Consumable", "Chemicals - Solid"])
    wb.save(excel_path)

    imported_count, _ = import_items_from_excel(str(excel_path), editor_name="tester")
    assert imported_count == 1

    rows = db.execute_query(
        """
        SELECT c.name as category_name
        FROM Items i
        JOIN Categories c ON c.id = i.category_id
        WHERE i.name = ?
        """,
        ("AliasChem",),
    )
    assert rows
    assert rows[0]["category_name"] == "Chemicals-Solid"
